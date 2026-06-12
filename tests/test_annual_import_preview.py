"""End-to-end tests for the annual budget import preview/confirm/cancel flow."""
from io import BytesIO

from openpyxl import load_workbook


def _seed(app, uid):
    """Seed a minimal budget: 1 income item, 1 linked-ISA item, 1 account."""
    from app.models import get_connection, create_budget_item, upsert_budget_entry
    with app.app_context():
        with get_connection() as conn:
            conn.execute(
                "INSERT INTO budget_sections (user_id, key, label, sort_order) VALUES (?, 'income', 'Income', 1)",
                (uid,),
            )
            conn.execute(
                "INSERT INTO budget_sections (user_id, key, label, sort_order) VALUES (?, 'inv', 'Investments', 2)",
                (uid,),
            )
            conn.execute(
                """INSERT INTO accounts (user_id, name, wrapper_type, current_value, monthly_contribution, is_active)
                   VALUES (?, 'My ISA', 'Stocks & Shares ISA', 5000, 333, 1)""",
                (uid,),
            )
            conn.commit()
            aid = conn.execute("SELECT id FROM accounts WHERE user_id=?", (uid,)).fetchone()["id"]

        create_budget_item(
            {"name": "Salary", "section": "income", "default_amount": 3500,
             "linked_account_id": None, "notes": "", "sort_order": 1}, uid)
        isa_id = create_budget_item(
            {"name": "ISA", "section": "inv", "default_amount": 333,
             "linked_account_id": aid, "notes": "", "sort_order": 1}, uid)
        upsert_budget_entry("2026-04", isa_id, 333, uid)
        return aid, isa_id


def _export_and_edit(auth_client, edits):
    """Download annual export, apply {"Sheet Name": {"ISA": 450, ...}} edits, return file bytes."""
    r = auth_client.get("/budget/annual-export.xlsx")
    wb = load_workbook(BytesIO(r.data))
    for sheet_name, row_edits in edits.items():
        ws = wb[sheet_name]
        for r_idx in range(4, ws.max_row + 1):
            name = ws.cell(r_idx, 2).value
            if isinstance(name, str) and name in row_edits:
                ws.cell(r_idx, 4).value = row_edits[name]
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def test_annual_export_uses_lifetime_isa_copy(auth_client, app, make_user):
    uid, _, _ = make_user(username="annual-export-lifetime-isa", password="password123")
    _seed(app, uid)

    resp = auth_client.get("/budget/annual-export.xlsx")
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.data), data_only=True)

    guide_values = [cell.value for row in wb["Guide"].iter_rows() for cell in row if isinstance(cell.value, str)]
    tracking_values = [cell.value for row in wb["Investment Tracking"].iter_rows() for cell in row if isinstance(cell.value, str)]

    assert "- Lifetime ISA bonus: 25% bonus on Lifetime ISA personal contributions until the £4,000/year cap (bonus does not count toward ISA allowance)." in guide_values
    assert "- LISA bonus: 25% bonus on LISA personal contributions until the £4,000/year cap (bonus does not count toward ISA allowance)." not in guide_values
    assert "  of which Lifetime ISA" in tracking_values
    assert "  of which LISA" not in tracking_values
    assert "Note: Lifetime ISA personal contributions count toward the overall ISA £20k allowance. The 25% Lifetime ISA bonus does not." in tracking_values
    assert "Note: LISA personal contributions count toward the overall ISA £20k allowance. The 25% LISA bonus does not." not in tracking_values
    assert "Planned Lifetime ISA bonus" in tracking_values
    assert "Logged Lifetime ISA bonus" in tracking_values
    assert tracking_values.count("Lifetime ISA bonus") == 3
    assert "Planned LISA bonus" not in tracking_values
    assert "Logged LISA bonus" not in tracking_values
    assert "LISA bonus" not in tracking_values


def test_annual_import_preview_does_not_write_db(app, auth_client, make_user):
    from app.models import get_connection
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {"Jun 2026": {"ISA": 450}})
    resp = auth_client.post("/budget/annual-import",
                            data={"file": (buf, "t.xlsx")},
                            content_type="multipart/form-data")
    assert resp.status_code == 200
    assert b"Preview" in resp.data
    assert b"June 2026" in resp.data
    assert b"450" in resp.data
    assert b"Budget Setup" in resp.data
    assert b"Edit Budget" not in resp.data

    with app.app_context():
        with get_connection() as conn:
            # No June entry written yet; no overrides yet
            row = conn.execute(
                "SELECT COUNT(*) c FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-06'", (uid,)).fetchone()
            assert row["c"] == 0
            assert conn.execute("SELECT COUNT(*) c FROM contribution_overrides").fetchone()["c"] == 0


def test_annual_import_confirm_writes_db_and_syncs_overrides(app, auth_client, make_user):
    from app.models import get_connection
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {"Jun 2026": {"ISA": 450}, "Oct 2026": {"Salary": 3700}})
    auth_client.post("/budget/annual-import",
                     data={"file": (buf, "t.xlsx")},
                     content_type="multipart/form-data")
    resp = auth_client.post("/budget/annual-import/confirm")
    assert resp.status_code == 302

    with app.app_context():
        with get_connection() as conn:
            jun = conn.execute(
                "SELECT amount FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-06' AND bi.name='ISA'", (uid,)).fetchone()
            oct_sal = conn.execute(
                "SELECT amount FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-10' AND bi.name='Salary'", (uid,)).fetchone()
            override = conn.execute(
                "SELECT override_amount FROM contribution_overrides WHERE from_month='2026-06'").fetchone()
            assert jun["amount"] == 450
            assert oct_sal["amount"] == 3700
            assert override["override_amount"] == 450  # linked item → override synced


def test_annual_import_empty_diff_redirects_without_preview(app, auth_client, make_user):
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {})  # no edits → matches current state
    resp = auth_client.post("/budget/annual-import",
                            data={"file": (buf, "t.xlsx")},
                            content_type="multipart/form-data",
                            follow_redirects=False)
    assert resp.status_code == 302
    assert "/budget/" in resp.headers["Location"]


def test_annual_import_cancel_discards_staged_changes(app, auth_client, make_user):
    from app.models import get_connection
    uid, _, _ = make_user()
    _seed(app, uid)

    buf = _export_and_edit(auth_client, {"Jun 2026": {"ISA": 450}})
    auth_client.post("/budget/annual-import",
                     data={"file": (buf, "t.xlsx")},
                     content_type="multipart/form-data")
    resp = auth_client.post("/budget/annual-import/cancel", follow_redirects=False)
    assert resp.status_code == 302

    with app.app_context():
        with get_connection() as conn:
            row = conn.execute(
                "SELECT COUNT(*) c FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.month_key='2026-06'", (uid,)).fetchone()
            assert row["c"] == 0


def test_annual_import_stale_confirm_redirects_gracefully(app, auth_client, make_user):
    uid, _, _ = make_user()
    _seed(app, uid)

    resp = auth_client.post("/budget/annual-import/confirm", follow_redirects=False)
    assert resp.status_code == 302


def test_annual_import_staging_survives_large_payload(app, auth_client, make_user):
    """The disk-backed staging must handle a diff larger than the 4KB signed-
    cookie limit. We simulate this by writing many budget items and editing
    every month for each one."""
    from app.models import get_connection, create_budget_item
    uid, _, _ = make_user()
    _seed(app, uid)

    with app.app_context():
        # ~50 items × 12 months = 600 potential rows
        for i in range(50):
            create_budget_item(
                {"name": f"Item {i}", "section": "inv", "default_amount": 100 + i,
                 "linked_account_id": None, "notes": "", "sort_order": i + 10},
                uid,
            )

    # Edit every item in every month to a different value
    r = auth_client.get("/budget/annual-export.xlsx")
    wb = load_workbook(BytesIO(r.data))
    for sheet in wb.sheetnames:
        if sheet in ("Summary", "Investment Tracking"):
            continue
        ws = wb[sheet]
        for r_idx in range(4, ws.max_row + 1):
            name = ws.cell(r_idx, 2).value
            if isinstance(name, str) and name.startswith("Item "):
                ws.cell(r_idx, 4).value = 999  # everything → 999
    buf = BytesIO(); wb.save(buf); buf.seek(0)

    resp = auth_client.post("/budget/annual-import",
                            data={"file": (buf, "t.xlsx")},
                            content_type="multipart/form-data")
    assert resp.status_code == 200
    # Preview renders even though the diff is big (would have broken with
    # cookie-only staging: ~50 items × 12 months × ~60 bytes > 4KB)

    # Confirm should succeed — i.e. the staged payload is still readable
    resp2 = auth_client.post("/budget/annual-import/confirm", follow_redirects=False)
    assert resp2.status_code == 302
    with app.app_context():
        with get_connection() as conn:
            count = conn.execute(
                "SELECT COUNT(*) c FROM budget_entries be JOIN budget_items bi ON bi.id=be.budget_item_id "
                "WHERE bi.user_id=? AND be.amount=999", (uid,)).fetchone()["c"]
        assert count >= 50 * 12 * 0.9  # allow a bit of slack; round-trip noise excluded


def test_annual_import_staging_rejects_cross_user_token(app, auth_client, client, make_user):
    """A tampered session cookie pointing at another user's token must not
    allow confirming their staged diff."""
    from app.services.import_staging import write_staged

    alice_uid, _, _ = make_user(username="alice")
    bob_uid, _, _ = make_user(username="bob")
    _seed(app, alice_uid)

    # Write a staging file that claims it belongs to Alice
    with app.app_context():
        alice_token = write_staged(app, {
            "user_id": alice_uid,
            "changes": [{"item_id": 999, "month_key": "2026-07", "new": 5000, "linked": False}],
        })

    # Log in as Bob, plant Alice's token in Bob's session, try to confirm
    client.post("/login", data={"username": "bob", "password": "testpass123"},
                follow_redirects=False)
    with client.session_transaction() as s:
        s["budget_annual_import"] = {"token": alice_token, "user_id": alice_uid}

    resp = client.post("/budget/annual-import/confirm", follow_redirects=False)
    # Staged payload belongs to Alice, session claims Bob → mismatch, reject
    assert resp.status_code == 302
    with app.app_context():
        from app.models import get_connection
        with get_connection() as conn:
            # No entry was written for Bob (or anyone)
            count = conn.execute(
                "SELECT COUNT(*) c FROM budget_entries WHERE amount=5000").fetchone()["c"]
        assert count == 0
