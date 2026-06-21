from io import BytesIO

from openpyxl import load_workbook

from app.models import create_account, get_connection


def _account_payload(name="ISA", wrapper_type="Stocks & Shares ISA", value=1000.0, monthly=0.0):
    return {
        "name": name,
        "provider": "Provider",
        "wrapper_type": wrapper_type,
        "category": "Investment",
        "tags": "",
        "current_value": value,
        "monthly_contribution": monthly,
        "pension_contribution_day": 0,
        "goal_value": None,
        "valuation_mode": "manual",
        "growth_mode": "default",
        "growth_rate_override": None,
        "owner": "Janusz",
        "is_active": 1,
        "notes": "",
        "last_updated": "2026-05-25",
        "employer_contribution": 0,
        "contribution_method": "standard",
        "annual_fee_pct": 0,
        "platform_fee_pct": 0,
        "platform_fee_flat": 0,
        "platform_fee_cap": 0,
        "fund_fee_pct": 0,
        "contribution_fee_pct": 0,
        "uninvested_cash": 0,
        "cash_interest_rate": 0,
        "interest_payment_day": 0,
    }


def _login(client, username, password):
    client.post("/login", data={"username": username, "password": password}, follow_redirects=False)


def _workbook_from_response(response):
    assert response.status_code == 200
    return load_workbook(BytesIO(response.data), data_only=True)


def test_performance_export_keeps_zero_snapshot_message_for_selected_account(app, client, make_user):
    uid, username, password = make_user(username="perf-export-zero", password="password123")
    with app.app_context():
        account_id = create_account(_account_payload(), uid)

    _login(client, username, password)
    workbook = _workbook_from_response(client.get(f"/performance/export.xlsx?account_id={account_id}"))

    detail = workbook["ISA (Monthly)"]
    assert detail["A4"].value == "Not enough data yet (need at least two monthly snapshots)."
    assert detail["A5"].value is None


def test_performance_export_shows_first_snapshot_as_initial_funding_for_portfolio_and_account(app, client, make_user):
    uid, username, password = make_user(username="perf-export-first-baseline", password="password123")
    with app.app_context():
        account_id = create_account(_account_payload(value=6000, monthly=150), uid)
        with get_connection() as conn:
            conn.execute(
                "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES ('2026-05-01', ?, 6000, '2026-05')",
                (account_id,),
            )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    portfolio = workbook["Portfolio (Monthly)"]
    assert portfolio["A4"].value == "Month"
    assert portfolio["A5"].value == "May 2026"
    assert portfolio["B5"].value == 0
    assert portfolio["C5"].value == 6000
    assert portfolio["D5"].value == 0
    assert portfolio["E5"].value == 0
    assert portfolio["F5"].value == 6000
    assert "Your first baseline is saved" not in str(portfolio["A4"].value)

    account = workbook["ISA (Monthly)"]
    assert account["A4"].value == "Month"
    assert account["A5"].value == "May 2026"
    assert account["B5"].value == 0
    assert account["C5"].value == 6000
    assert account["D5"].value == 0
    assert account["E5"].value == 0
    assert account["F5"].value == 6000
    assert "Your first baseline is saved" not in str(account["A4"].value)



def test_performance_export_shows_initial_funding_rows_for_one_snapshot_accounts(app, client, make_user):
    uid, username, password = make_user(username="perf-export-initial-funding", password="password123")
    with app.app_context():
        premium_id = create_account(_account_payload(name="Premium Bonds", wrapper_type="Premium Bonds", value=800, monthly=0), uid)
        my_sipp_id = create_account(_account_payload(name="My SIPP", wrapper_type="SIPP", value=206.36, monthly=0), uid)
        with get_connection() as conn:
            for account_id, balance in [(premium_id, 800), (my_sipp_id, 206.36)]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES ('2026-06-01', ?, ?, '2026-06')",
                    (account_id, balance),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    summary = workbook["Summary"]
    imported = {summary.cell(row=r, column=1).value: summary.cell(row=r, column=8).value for r in range(5, summary.max_row + 1)}
    assert imported["Portfolio"] == 1006.36
    assert imported["Premium Bonds"] == 800
    assert imported["My SIPP"] == 206.36

    premium = workbook["Premium Bonds (Monthly)"]
    assert premium["A4"].value == "Month"
    assert premium["A5"].value == "Jun 2026"
    assert premium["B5"].value == 0
    assert premium["C5"].value == 800
    assert premium["D5"].value == 0
    assert premium["E5"].value == 0
    assert premium["F5"].value == 800

    my_sipp = workbook["My SIPP (Monthly)"]
    assert my_sipp["A4"].value == "Month"
    assert my_sipp["A5"].value == "Jun 2026"
    assert my_sipp["B5"].value == 0
    assert my_sipp["C5"].value == 206.36
    assert my_sipp["D5"].value == 0
    assert my_sipp["E5"].value == 0
    assert my_sipp["F5"].value == 206.36


def test_performance_export_uses_live_current_month_values_for_end_balances(app, client, make_user):
    uid, username, password = make_user(username="perf-export-live-current", password="password123")
    with app.app_context():
        stale_id = create_account(_account_payload(name="Stale ISA", value=2500, monthly=100), uid)
        new_id = create_account(_account_payload(name="New SIPP", wrapper_type="SIPP", value=206, monthly=250), uid)
        with get_connection() as conn:
            conn.execute(
                "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES ('2026-05-01', ?, 2000, '2026-05')",
                (stale_id,),
            )
            conn.execute(
                "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES ('2026-06-01', ?, 2000, '2026-06')",
                (stale_id,),
            )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    summary = workbook["Summary"]
    rows = {summary.cell(row=r, column=1).value: summary.cell(row=r, column=11).value for r in range(5, summary.max_row + 1)}

    assert rows["Portfolio"] == 2706
    assert rows["Stale ISA"] == 2500
    assert rows["New SIPP"] == 206
    assert "New SIPP (Monthly)" in workbook.sheetnames


def test_performance_export_uses_cash_flow_events_for_monthly_attribution(app, client, make_user):
    uid, username, password = make_user(username="perf-export-cashflows", password="password123")
    with app.app_context():
        cash_id = create_account(_account_payload(name="Cash ISA", wrapper_type="Cash ISA", value=1600, monthly=1000), uid)
        sipp_id = create_account(_account_payload(name="SIPP", wrapper_type="SIPP", value=1400, monthly=0), uid)
        with get_connection() as conn:
            for account_id, month_key, balance in [
                (cash_id, "2026-05", 3000),
                (cash_id, "2026-06", 1600),
                (sipp_id, "2026-06", 1400),
            ]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.execute(
                """
                INSERT INTO cash_flow_events (user_id, account_id, event_date, amount, kind, counterparty_account_id, note, created_at)
                VALUES (?, ?, '2026-06-15', -1400, 'transfer_out', ?, 'Moved to SIPP', '2026-06-15T00:00:00+00:00')
                """,
                (uid, cash_id, sipp_id),
            )
            conn.execute(
                """
                INSERT INTO cash_flow_events (user_id, account_id, event_date, amount, kind, counterparty_account_id, note, created_at)
                VALUES (?, ?, '2026-06-15', 1400, 'deposit', ?, 'Transfer from Cash ISA', '2026-06-15T00:00:00+00:00')
                """,
                (uid, sipp_id, cash_id),
            )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    summary = workbook["Summary"]
    rows = {summary.cell(row=r, column=1).value: summary.cell(row=r, column=7).value for r in range(5, summary.max_row + 1)}
    assert rows["Portfolio"] == 0
    assert rows["Cash ISA"] == -1400
    assert rows["SIPP"] == 1400

    portfolio = workbook["Portfolio (Monthly)"]
    assert portfolio["C5"].value == 0
    assert portfolio["D5"].value == 0

    cash = workbook["Cash ISA (Monthly)"]
    assert cash["C5"].value == 0
    assert cash["D5"].value == -1400
    assert cash["E5"].value == 0


def test_performance_export_shows_actual_first_month_contribution(app, client, make_user):
    uid, username, password = make_user(username="perf-export-first-cashflow", password="password123")
    with app.app_context():
        account_id = create_account(_account_payload(name="My SIPP", wrapper_type="SIPP", value=206, monthly=0), uid)
        with get_connection() as conn:
            conn.execute(
                "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES ('2026-06-01', ?, 206.36, '2026-06')",
                (account_id,),
            )
            conn.execute(
                """
                INSERT INTO cash_flow_events (user_id, account_id, event_date, amount, kind, note, created_at)
                VALUES (?, ?, '2026-06-01', 200, 'deposit', 'Opened SIPP', '2026-06-01T00:00:00+00:00')
                """,
                (uid, account_id),
            )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    account = workbook["My SIPP (Monthly)"]
    assert account["A4"].value == "Month"
    assert account["A5"].value == "Jun 2026"
    assert account["B5"].value == 6
    assert account["C5"].value == 0
    assert account["D5"].value == 200
    assert account["E5"].value == 0
    assert account["F5"].value == 206


def test_performance_export_does_not_count_unconfirmed_current_month_plan_as_actual_contribution(app, client, make_user):
    uid, username, password = make_user(username="perf-export-current-plan", password="password123")
    with app.app_context():
        account_id = create_account(_account_payload(name="Stocks & Shares ISA", value=7774.07, monthly=500), uid)
        with get_connection() as conn:
            for month_key, balance in [("2026-05", 7609.35), ("2026-06", 7774.07)]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    account = workbook["Stocks & Shares ISA (Monthly)"]
    assert account["A5"].value == "Jun 2026"
    assert account["B5"].value == 7609.35
    assert account["C5"].value == 0
    assert account["D5"].value == 0
    assert account["E5"].value == 164.72
    assert account["F5"].value == 7774.07


def test_performance_export_treats_cash_isa_untracked_balance_drops_as_cash_movements(app, client, make_user):
    uid, username, password = make_user(username="perf-export-cash-drop", password="password123")
    with app.app_context():
        account_id = create_account(_account_payload(name="Cash ISA", wrapper_type="Cash ISA", value=1597.51, monthly=0), uid)
        with get_connection() as conn:
            for month_key, balance in [("2026-05", 1947.86), ("2026-06", 1597.51)]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    account = workbook["Cash ISA (Monthly)"]
    assert account["A5"].value == "Jun 2026"
    assert account["B5"].value == 1947.86
    assert account["C5"].value == 0
    assert account["D5"].value == -350.35
    assert account["E5"].value == 0
    assert account["F5"].value == 1597.51
    assert account["G5"].value == 0


def test_performance_export_counts_regular_contribution_after_account_specific_pension_day(monkeypatch, app, client, make_user):
    from datetime import date as real_date, datetime as real_datetime
    import app.models.planning_snapshots as planning_snapshots
    import app.routes.export as export_route

    class FixedDatetime(real_datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 21, 12, 0, 0, tzinfo=tz)

    class FixedDate(real_date):
        @classmethod
        def today(cls):
            return cls(2026, 6, 21)

    monkeypatch.setattr(planning_snapshots, "datetime", FixedDatetime)
    monkeypatch.setattr(planning_snapshots, "date", FixedDate)
    monkeypatch.setattr(export_route, "date", FixedDate)

    uid, username, password = make_user(username="perf-export-workplace-pension-day", password="password123")
    with app.app_context():
        with get_connection() as conn:
            conn.execute("UPDATE assumptions SET salary_day = 28 WHERE user_id = ?", (uid,))
            conn.commit()
        account_id = create_account(
            {
                **_account_payload(name="Workplace Pension", wrapper_type="Workplace Pension", value=3445.27, monthly=333.33),
                "category": "Pension",
                "contribution_method": "salary_sacrifice",
                "pension_contribution_day": 19,
            },
            uid,
        )
        with get_connection() as conn:
            for month_key, balance in [("2026-05", 3031.02), ("2026-06", 3445.27)]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    account = workbook["Workplace Pension (Monthly)"]
    assert account["A5"].value == "Jun 2026"
    assert account["C5"].value == 0
    assert account["D5"].value == 333.33
    assert account["E5"].value == 80.92


def test_performance_export_labels_cash_isa_with_cash_interest_and_not_global_growth(app, client, make_user):
    uid, username, password = make_user(username="perf-export-cash-rate", password="password123")
    with app.app_context():
        account_id = create_account(
            {
                **_account_payload(name="Cash ISA", wrapper_type="Cash ISA", value=1597.51, monthly=0),
                "cash_interest_rate": 0.036,
            },
            uid,
        )
        with get_connection() as conn:
            for month_key, balance in [("2026-04", 2600), ("2026-05", 1947.86), ("2026-06", 1597.51)]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    summary = workbook["Summary"]
    cash_row = next(r for r in range(5, summary.max_row + 1) if summary.cell(row=r, column=1).value == "Cash ISA")
    total_return = float(summary.cell(row=cash_row, column=5).value or 0)
    total_gain = float(summary.cell(row=cash_row, column=9).value or 0)
    assert 0 < total_return < 1
    assert 0 < total_gain < 20

    account = workbook["Cash ISA (Monthly)"]
    assert account["A2"].value == "Cash interest rate: 3.6%"
    assert account["E4"].value == "Interest / Cash gain"
    assert 0 < float(account["E5"].value or 0) < 10
    assert 0 < float(account["E6"].value or 0) < 10


def test_performance_export_labels_premium_bonds_with_expected_prize_rate(app, client, make_user):
    uid, username, password = make_user(username="perf-export-premium-bonds-rate", password="password123")
    with app.app_context():
        account_id = create_account(
            {
                **_account_payload(name="Premium Bonds", wrapper_type="Premium Bonds", value=800, monthly=0),
                "growth_mode": "custom",
                "growth_rate_override": 0.033,
            },
            uid,
        )
        with get_connection() as conn:
            for month_key, balance in [("2026-05", 800), ("2026-06", 800)]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    account = workbook["Premium Bonds (Monthly)"]
    assert account["A2"].value == "Expected prize rate: 3.3%"
    assert account["E4"].value == "Prize gain"


def test_performance_export_counts_regular_contribution_after_salary_day(monkeypatch, app, client, make_user):
    from datetime import date as real_date, datetime as real_datetime
    import app.models.planning_snapshots as planning_snapshots
    import app.routes.export as export_route

    class FixedDatetime(real_datetime):
        @classmethod
        def now(cls, tz=None):
            return cls(2026, 6, 29, 12, 0, 0, tzinfo=tz)

    class FixedDate(real_date):
        @classmethod
        def today(cls):
            return cls(2026, 6, 29)

    monkeypatch.setattr(planning_snapshots, "datetime", FixedDatetime)
    monkeypatch.setattr(export_route, "date", FixedDate)

    uid, username, password = make_user(username="perf-export-after-salary-day", password="password123")
    with app.app_context():
        with get_connection() as conn:
            conn.execute("UPDATE assumptions SET salary_day = 28 WHERE user_id = ?", (uid,))
            conn.commit()
        account_id = create_account(_account_payload(name="Stocks & Shares ISA", value=8109.35, monthly=500), uid)
        with get_connection() as conn:
            for month_key, balance in [("2026-05", 7609.35), ("2026-06", 8109.35)]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx"))

    account = workbook["Stocks & Shares ISA (Monthly)"]
    assert account["A5"].value == "Jun 2026"
    assert account["C5"].value == 0
    assert account["D5"].value == 500
    assert account["E5"].value == 0


def test_performance_export_filters_to_requested_historical_window(app, client, make_user):
    uid, username, password = make_user(username="perf-export-window", password="password123")
    with app.app_context():
        account_id = create_account(_account_payload(value=1500, monthly=100), uid)
        with get_connection() as conn:
            for month_key, balance in [
                ("2026-01", 1000),
                ("2026-02", 1100),
                ("2026-03", 1200),
                ("2026-04", 1350),
                ("2026-05", 1500),
            ]:
                conn.execute(
                    "INSERT INTO monthly_snapshots (snapshot_date, account_id, balance, month_key) VALUES (?, ?, ?, ?)",
                    (f"{month_key}-01", account_id, balance, month_key),
                )
            conn.commit()

    _login(client, username, password)
    workbook = _workbook_from_response(client.get("/performance/export.xlsx?period=1M"))

    summary = workbook["Summary"]
    assert summary["B5"].value == "Apr 2026"
    assert summary["C5"].value == "May 2026"
    assert summary["D5"].value == 1

    portfolio = workbook["Portfolio (Monthly)"]
    assert portfolio["A5"].value == "May 2026"
    assert portfolio["A6"].value is None

    account = workbook["ISA (Monthly)"]
    assert account["A5"].value == "May 2026"
    assert account["A6"].value is None
