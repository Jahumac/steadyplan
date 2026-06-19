from datetime import date as real_date
from io import BytesIO

from openpyxl import load_workbook

from app.models import create_account, create_contribution_override, create_temporary_contribution_plan, fetch_assumptions, update_assumptions
from app.calculations import projection_start_month_key


def _account(name, wrapper_type, value=0, monthly=0, **overrides):
    payload = {
        "id": None,
        "name": name,
        "provider": "Provider",
        "wrapper_type": wrapper_type,
        "category": "Investment",
        "tags": "",
        "current_value": value,
        "monthly_contribution": monthly,
        "pension_contribution_day": 0,
        "goal_value": None,
        "valuation_mode": "manual",
        "growth_mode": "default",
        "growth_rate_override": None,
        "owner": "Janusz",
        "is_active": 1,
        "notes": "",
        "last_updated": "2026-05-25",
        "employer_contribution": 0,
        "contribution_method": "standard",
        "annual_fee_pct": 0,
        "platform_fee_pct": 0,
        "platform_fee_flat": 0,
        "platform_fee_cap": 0,
        "fund_fee_pct": 0,
        "contribution_fee_pct": 0,
        "uninvested_cash": 0,
        "cash_interest_rate": 0,
        "interest_payment_day": 0,
    }
    payload.update(overrides)
    return payload


def _login(client, username, password):
    client.post("/login", data={"username": username, "password": password}, follow_redirects=False)


def _workbook_from_response(resp):
    assert resp.status_code == 200
    return load_workbook(BytesIO(resp.data), data_only=True)


def test_projections_export_explains_assumptions_schedule_and_access(app, client, make_user):
    uid, username, password = make_user(username="projection-export", password="password123")
    with app.app_context():
        assumptions = dict(fetch_assumptions(uid))
        assumptions.update({
            "annual_growth_rate": 0.05,
            "retirement_age": 62,
            "date_of_birth": "1983-05-25",
            "salary_day": 25,
        })
        update_assumptions(assumptions, uid)
        isa_id = create_account(_account("ISA", "Stocks & Shares ISA", 10000, 500), uid)
        create_account(_account("Cash ISA", "Cash ISA", 2000, 0), uid)
        create_account(_account("SIPP", "SIPP", 50000, 100, contribution_method="relief_at_source"), uid)
        create_account(_account("LISA", "Lifetime ISA", 1000, 0), uid)
        create_contribution_override({
            "account_id": isa_id,
            "from_month": "2028-11",
            "to_month": "9999-12",
            "override_amount": 750,
            "reason": "Future increase",
        }, uid)

    _login(client, username, password)
    wb = _workbook_from_response(client.get("/projections/export.xlsx"))

    assert "Scenario Estimate Assumptions" in wb.sheetnames
    assert "Assumptions" not in wb.sheetnames
    assert "Contribution Schedule" in wb.sheetnames

    summary = wb["Summary"]
    assert [summary.cell(4, c).value for c in range(1, 6)] == [
        "Account", "Current Value", "You pay monthly", "Into pots monthly", "Scenario estimate at retirement"
    ]
    summary_values = [cell.value for row in summary.iter_rows() for cell in row]
    assert "Cash-accessible, invested-accessible, restricted, and locked-for-later money" in summary_values
    assert "Accessible now vs locked" not in summary_values
    assert "Accessible vs locked" not in summary_values
    assert "Cash accessible" in summary_values
    assert "Invested accessible" in summary_values
    assert "Accessible" not in summary_values
    assert "Locked for later" in summary_values
    assert "Values are nominal scenario estimates before inflation unless stated otherwise." in summary_values
    assert "Values are nominal projections before inflation unless stated otherwise." not in summary_values
    assert "Projected at Retirement" not in summary_values
    assert "Projected at retirement" not in summary_values

    assumptions_sheet = wb["Scenario Estimate Assumptions"]
    assumption_values = [cell.value for row in assumptions_sheet.iter_rows() for cell in row]
    assert "Scenario estimate start month" in assumption_values
    assert "Inflation treatment" in assumption_values
    assert "Nominal" in assumption_values
    assert "Target age used for this scenario estimate." in assumption_values
    assert "Target age used for this projection." not in assumption_values
    assert "Projection start month" not in assumption_values
    assert "First future contribution month considered by scenario estimates." in assumption_values
    assert "First future contribution month considered by projections." not in assumption_values

    workbook_values = [cell.value for sheet in wb.worksheets for row in sheet.iter_rows() for cell in row if cell.value is not None]
    assert "SteadyPlan — Retirement Scenario Estimates" in workbook_values
    assert "SteadyPlan — Scenario Estimate Assumptions" in workbook_values
    assert "SteadyPlan — Year-by-Year Scenario Estimate" in workbook_values
    assert "SteadyPlan — Monthly Scenario Estimate" in workbook_values
    assert "Scenario estimate total" in workbook_values
    assert "Scenario estimate value" in workbook_values
    assert "SteadyPlan — Retirement Projections" not in workbook_values
    assert "SteadyPlan — Projection Assumptions" not in workbook_values
    assert "SteadyPlan — Year-by-Year Projection" not in workbook_values
    assert "SteadyPlan — Monthly Projection" not in workbook_values
    assert "Projected Total" not in workbook_values
    assert "Projected Value" not in workbook_values

    schedule = wb["Contribution Schedule"]
    schedule_rows = [tuple(cell.value for cell in row) for row in schedule.iter_rows()]
    assert ("ISA", "Stocks & Shares ISA", "2028-11", "Ongoing", 750, 750, "Future increase") in schedule_rows
    assert ("ISA", "Stocks & Shares ISA", "2028-11", "9999-12", 750, 750, "Future increase") not in schedule_rows


def test_lifetime_isa_one_off_override_export_includes_full_bonus(app, client, make_user):
    uid, username, password = make_user(username="projection-export-lisa-one-off", password="password123")
    with app.app_context():
        assumptions = dict(fetch_assumptions(uid))
        assumptions.update({
            "annual_growth_rate": 0.0,
            "retirement_age": 45,
            "date_of_birth": "1983-05-25",
            "salary_day": 25,
        })
        update_assumptions(assumptions, uid)
        start_month = projection_start_month_key(assumptions)
        lisa_id = create_account(_account("Lifetime ISA", "Lifetime ISA", 1000, 0, growth_mode="custom", growth_rate_override=0), uid)
        create_contribution_override({
            "account_id": lisa_id,
            "from_month": start_month,
            "to_month": start_month,
            "override_amount": 4000,
            "reason": "LISA lump sum",
        }, uid)

    _login(client, username, password)
    wb = _workbook_from_response(client.get("/projections/export.xlsx"))

    schedule_rows = [tuple(cell.value for cell in row) for row in wb["Contribution Schedule"].iter_rows()]
    assert ("Lifetime ISA", "Lifetime ISA", start_month, start_month, 4000, 5000, "LISA lump sum") in schedule_rows

    ws = wb["Lifetime ISA"]
    headers = None
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[:3] == ("Age", "Year", "Scenario estimate value"):
            headers = row
            header_row = idx
            break
    assert headers is not None
    assert header_row is not None
    you_pay_col = headers.index("You pay (yr)")
    into_pot_col = headers.index("Into pot (yr)")
    first_year = next(row for row in ws.iter_rows(min_row=header_row + 1, values_only=True) if row[you_pay_col] == 4000)
    assert first_year[into_pot_col] == 5000


def test_lifetime_isa_yearly_export_uses_calendar_year_distribution(app, client, make_user, monkeypatch):
    from app import calculations
    from app.routes import export as export_routes

    class FakeDate(real_date):
        @classmethod
        def today(cls):
            return cls(2026, 6, 18)

    monkeypatch.setattr(calculations, "date", FakeDate)
    monkeypatch.setattr(export_routes, "date", FakeDate)

    uid, username, password = make_user(username="projection-export-lisa-calendar-year", password="password123")
    with app.app_context():
        assumptions = dict(fetch_assumptions(uid))
        assumptions.update({
            "annual_growth_rate": 0.0,
            "retirement_age": 46,
            "date_of_birth": "1982-12-18",
            "salary_day": 28,
        })
        update_assumptions(assumptions, uid)
        lisa_id = create_account(_account("Lifetime ISA", "Lifetime ISA", 277, 0, growth_mode="custom", growth_rate_override=0), uid)
        create_contribution_override({
            "account_id": lisa_id,
            "from_month": "2026-12",
            "to_month": "2027-07",
            "override_amount": 1000,
            "reason": "Tax-year LISA fill",
        }, uid)

    _login(client, username, password)
    wb = _workbook_from_response(client.get("/projections/export.xlsx"))

    ws = wb["Lifetime ISA"]
    headers = None
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[:6] == ("Age", "Year", "Scenario estimate value", "Growth", "You pay (yr)", "Into pot (yr)"):
            headers = row
            header_row = idx
            break
    assert headers is not None
    assert header_row is not None

    data_rows = [row for row in ws.iter_rows(min_row=header_row + 1, values_only=True) if isinstance(row[1], (int, str))]
    row_today = next(row for row in data_rows if row[1] == "2026 (today)")
    row_2026 = next(row for row in data_rows if row[1] == 2026)
    row_2027 = next(row for row in data_rows if row[1] == 2027)

    assert row_today[4] == 0
    assert row_today[5] == 0
    assert row_2026[4] == 1000
    assert row_2026[5] == 1250
    assert row_2027[4] == 7000
    assert row_2027[5] == 8750

    total_sheet_rows = [tuple(cell.value for cell in row) for row in wb["Year by Year"].iter_rows()]
    assert (43, "2026 (today)", 277) in total_sheet_rows
    assert (44, 2026, 1527) in total_sheet_rows
    assert (45, 2027, 10277) in total_sheet_rows


def test_lifetime_isa_export_shows_next_planned_month_when_current_month_is_zero(app, client, make_user, monkeypatch):
    from app import calculations
    from app.routes import export as export_routes

    class FakeDate(real_date):
        @classmethod
        def today(cls):
            return cls(2026, 6, 18)

    monkeypatch.setattr(calculations, "date", FakeDate)
    monkeypatch.setattr(export_routes, "date", FakeDate)

    uid, username, password = make_user(username="projection-export-lisa-next-planned", password="password123")
    with app.app_context():
        assumptions = dict(fetch_assumptions(uid))
        assumptions.update({
            "annual_growth_rate": 0.07,
            "retirement_age": 60,
            "date_of_birth": "1982-12-18",
            "salary_day": 28,
        })
        update_assumptions(assumptions, uid)
        lisa_id = create_account(_account(
            "Lifetime ISA",
            "Lifetime ISA",
            277,
            0,
            platform_fee_pct=0.15,
            fund_fee_pct=0.13,
        ), uid)
        create_temporary_contribution_plan(uid, "Yearly LISA fill", [{
            "account_id": lisa_id,
            "from_month": "2026-12",
            "to_month": "2027-07",
            "override_amount": 1000,
        }])

    _login(client, username, password)
    wb = _workbook_from_response(client.get("/projections/export.xlsx"))

    ws = wb["Lifetime ISA"]
    summary_rows = [tuple(cell.value for cell in row[:2]) for row in ws.iter_rows(min_row=5, max_row=16)]
    assert ("You pay (monthly)", 0) in summary_rows
    assert ("Next planned month", "2026-12") in summary_rows
    assert ("You pay (next planned month)", 1000) in summary_rows
    assert ("Total into pot (next planned month)", 1250) in summary_rows


def test_premium_bonds_cap_is_not_reported_as_negative_growth(app, client, make_user):
    uid, username, password = make_user(username="projection-export-pb", password="password123")
    with app.app_context():
        assumptions = dict(fetch_assumptions(uid))
        assumptions.update({
            "annual_growth_rate": 0.05,
            "retirement_age": 45,
            "date_of_birth": "1983-05-25",
        })
        update_assumptions(assumptions, uid)
        create_account(_account("Premium Bonds", "Premium Bonds", 49900, 300), uid)

    _login(client, username, password)
    wb = _workbook_from_response(client.get("/projections/export.xlsx"))

    ws = wb["Premium Bonds"]
    headers = None
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row[:3] == ("Age", "Year", "Scenario estimate value"):
            headers = row
            header_row = idx
            break
    assert headers is not None
    assert "Cap adjustment" in headers

    growth_col = headers.index("Growth") + 1
    cap_col = headers.index("Cap adjustment") + 1
    data_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    numeric_growth = [row[growth_col - 1] for row in data_rows if isinstance(row[growth_col - 1], (int, float))]
    numeric_caps = [row[cap_col - 1] for row in data_rows if isinstance(row[cap_col - 1], (int, float))]
    assert all(value >= 0 for value in numeric_growth)
    assert any(value < 0 for value in numeric_caps)

    schedule_values = [cell.value for row in wb["Contribution Schedule"].iter_rows() for cell in row]
    assert "NS&I Premium Bonds balance is capped; overflow is not shown as account growth." in schedule_values
