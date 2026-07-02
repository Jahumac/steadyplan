from datetime import date, datetime
from io import BytesIO

from flask import Blueprint, current_app, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required

from app.calculations import (
    add_months_to_key,
    apply_pension_carry_forward,
    contribution_breakdown,
    is_pension_account,
    pension_allowance_limits,
)
from app.utils import optional_float, optional_int, valid_month_key

from app.models import (
    build_debt_card,
    compare_debt_payoff_strategies,
    create_budget_item,
    create_budget_section,
    create_temporary_contribution_plan,
    create_debt,
    delete_budget_item,
    delete_budget_items_by_section,
    delete_budget_section,
    delete_temporary_contribution_plan,
    delete_debt,
    fetch_account,
    fetch_all_accounts,
    fetch_all_debts,
    fetch_budget_entries,
    fetch_budget_item,
    fetch_budget_items,
    fetch_budget_sections,
    fetch_budget_trend,
    fetch_contribution_calendar,
    fetch_temporary_contribution_plans,
    fetch_all_active_overrides,
    fetch_assumptions,
    fetch_debt,
    fetch_months_with_budget_entries,
    fetch_pension_carry_forward,
    fetch_prior_month_budget_entries,
    update_account,
    update_budget_item,
    update_budget_section,
    update_debt,
    upsert_budget_entry,
    upsert_single_month_contribution_override,
)
from app.models.debts import amortisation_schedule, schedule_anchor
from app.models.accounts import PREMIUM_BONDS_MAX_BALANCE, is_premium_bonds_account
from app.services.import_staging import (
    delete_staged,
    read_staged,
    write_staged,
)

budget_bp = Blueprint("budget", __name__)

def _default_month_key():
    today = date.today()
    return f"{today.year}-{today.month:02d}"


def _row_value(row, key, default=None):
    try:
        value = row[key]
    except (KeyError, IndexError, TypeError):
        return default
    return default if value is None else value


def _linked_account_default_amount(item, account):
    return float(_row_value(account, "monthly_contribution", 0) or 0)


def _parse_calendar_entry_key(raw_key):
    account_text, _, _ = str(raw_key or "").partition(":")
    try:
        account_id = int(account_text)
    except (TypeError, ValueError):
        return None, None
    return account_id, "invested"


def _linked_account_budget_total(month_key, user_id, account_id, account=None, entry_overrides=None):
    account = account or fetch_account(account_id, user_id)
    if not account:
        return 0.0

    items = fetch_budget_items(user_id)
    entries = fetch_budget_entries(month_key, user_id)
    entry_map = {e["budget_item_id"]: e for e in entries}
    entry_overrides = entry_overrides or {}
    total = 0.0

    for item in items:
        if _row_value(item, "linked_account_id") != account_id:
            continue
        if item["id"] in entry_overrides:
            amount = float(entry_overrides[item["id"]] or 0)
        elif item["id"] in entry_map:
            amount = float(entry_map[item["id"]]["amount"] or 0)
        else:
            amount = _linked_account_default_amount(item, account)
        total += amount

    return total


def _debt_payoff_month_key(debt):
    """Return YYYY-MM when the debt clears (from today), or None if it never does."""
    from app.models.debts import debt_months_remaining
    from app.calculations import add_months_to_key
    balance = float(debt["current_balance"] or 0)
    payment = float(debt["monthly_payment"] or 0)
    apr = float(debt["apr"] or 0)
    months = debt_months_remaining(balance, payment, apr)
    if not months:
        return None
    today = date.today()
    return add_months_to_key(f"{today.year}-{today.month:02d}", months)


def _sync_linked_override(item_id, month_key, amount, user_id):
    """When a linked budget item's entry changes, reflect it to contribution_overrides
    so projections and account views pick up the per-month edit automatically.

    Does nothing for unlinked items. Silently no-ops if the account isn't owned by
    user_id (handled by upsert_single_month_contribution_override). This keeps the
    account's normal monthly_contribution unchanged, so future months naturally
    fall back to the default unless they have their own override.
    """
    item = fetch_budget_item(item_id, user_id)
    if not item:
        return
    linked = _row_value(item, "linked_account_id")
    if not linked:
        return

    account = fetch_account(int(linked), user_id)
    if not account:
        return
    desired = float(amount or 0)
    total = _linked_account_budget_total(
        month_key,
        user_id,
        int(linked),
        account=account,
        entry_overrides={int(item_id): desired},
    )
    upsert_single_month_contribution_override(
        int(linked), month_key, total, user_id, reason="from budget"
    )

    today_key = _default_month_key()
    if month_key < today_key:
        return

    try:
        current = float(account.get("monthly_contribution") or 0)
    except (TypeError, ValueError):
        current = 0.0
    if abs(current - desired) < 0.005:
        return

    updated = dict(account)
    updated["monthly_contribution"] = desired
    update_account(updated, user_id)

def _month_label(month_key):
    return datetime.strptime(month_key, "%Y-%m").strftime("%B %Y")


def _default_contribution_calendar_range():
    today = date.today()
    start_year = today.year if today.month >= 4 else today.year - 1
    start_month = f"{start_year}-04"
    end_month = f"{start_year + 2}-03"
    return start_month, end_month


def _tax_year_label_for_month(month_key):
    year = int(month_key[:4])
    month = int(month_key[5:7])
    start_year = year if month >= 4 else year - 1
    return f"{start_year}/{str(start_year + 1)[-2:]}"


def _is_lifetime_isa_account(account):
    wrapper = (account.get("wrapper_type") or "").lower()
    return "lifetime isa" in wrapper or "lisa" in wrapper


def _is_isa_account(account):
    wrapper = (account.get("wrapper_type") or "").lower()
    return "isa" in wrapper or "lisa" in wrapper


def _is_cash_isa_account(account):
    wrapper = (account.get("wrapper_type") or "").lower()
    return "cash isa" in wrapper


def _is_stocks_and_shares_isa_account(account):
    wrapper = (account.get("wrapper_type") or "").lower()
    return "stocks" in wrapper and "isa" in wrapper


def _is_premium_bonds_account(account):
    return is_premium_bonds_account(account)


def _build_contribution_allowance_frame(calendar, assumptions, pension_carry_forward_entries=None):
    isa_allowance = float(assumptions["isa_allowance"]) if assumptions else 20000.0
    lisa_allowance = float(assumptions["lisa_allowance"]) if assumptions else 4000.0
    assumptions_map = dict(assumptions) if assumptions else {}
    pension_limits = apply_pension_carry_forward(
        pension_allowance_limits(assumptions_map),
        pension_carry_forward_entries or [],
    )
    pension_allowance = float(pension_limits.get("effective_allowance") or 0.0)
    pension_salary_cap = min(
        float(assumptions_map.get("annual_income") or 0.0),
        pension_allowance,
    ) if assumptions_map else 0.0
    premium_bonds_current_holding = sum(
        float(account.get("current_value") or 0.0)
        for account in calendar.get("accounts", [])
        if _is_premium_bonds_account(account)
    )
    by_tax_year = {}
    for month_key in calendar.get("months", []):
        tax_year = _tax_year_label_for_month(month_key)
        row = by_tax_year.setdefault(tax_year, {
            "tax_year": tax_year,
            "isa_planned": 0.0,
            "cash_isa_planned": 0.0,
            "stocks_and_shares_isa_planned": 0.0,
            "lisa_planned": 0.0,
            "pension_planned": 0.0,
        })
        for account in calendar.get("accounts", []):
            cell = next((c for c in account.get("months", []) if c.get("month_key") == month_key), None)
            if not cell:
                continue
            amount = float(cell["override_amount"] if cell.get("has_override") else cell.get("default_amount") or 0.0)
            if amount <= 0:
                continue
            if _is_lifetime_isa_account(account):
                row["isa_planned"] += amount
                row["lisa_planned"] += amount
            elif _is_cash_isa_account(account):
                row["isa_planned"] += amount
                row["cash_isa_planned"] += amount
            elif _is_stocks_and_shares_isa_account(account):
                row["isa_planned"] += amount
                row["stocks_and_shares_isa_planned"] += amount
            elif _is_isa_account(account):
                row["isa_planned"] += amount
            elif is_pension_account(account):
                adjusted = dict(account)
                adjusted["monthly_contribution"] = amount
                row["pension_planned"] += float(
                    contribution_breakdown(adjusted, assumptions).get("total_into_pot") or 0.0
                )

    rows = []
    for row in by_tax_year.values():
        row["isa_allowance"] = isa_allowance
        row["lisa_allowance"] = lisa_allowance
        row["pension_allowance"] = pension_allowance
        row["premium_bonds_holding"] = premium_bonds_current_holding
        row["premium_bonds_cap"] = PREMIUM_BONDS_MAX_BALANCE
        row["pension_personal_relief_limit"] = float(pension_limits.get("personal_relief_limit") or 0.0)
        row["pension_personal_tax_relief_cap"] = min(row["pension_personal_relief_limit"], row["pension_allowance"])
        row["pension_display_cap"] = pension_salary_cap or row["pension_personal_tax_relief_cap"] or row["pension_allowance"]
        row["pension_carry_forward_total"] = float(pension_limits.get("carry_forward_total") or 0.0)
        row["pension_mpaa_enabled"] = bool(pension_limits.get("mpaa_enabled"))
        row["isa_remaining"] = isa_allowance - row["isa_planned"]
        row["lisa_remaining"] = lisa_allowance - row["lisa_planned"]
        row["pension_remaining"] = pension_allowance - row["pension_planned"]
        row["isa_over"] = max(row["isa_planned"] - isa_allowance, 0.0)
        row["lisa_over"] = max(row["lisa_planned"] - lisa_allowance, 0.0)
        row["pension_over"] = max(row["pension_planned"] - pension_allowance, 0.0)
        row["has_warning"] = bool(row["isa_over"] or row["lisa_over"] or row["pension_over"])
        rows.append(row)
    return sorted(rows, key=lambda row: row["tax_year"])


def _build_monthly_data(month_key, user_id):
    db_sections = fetch_budget_sections(user_id)
    items = fetch_budget_items(user_id)
    entries = fetch_budget_entries(month_key, user_id)
    entry_map = {e["budget_item_id"]: e for e in entries}
    active_overrides = fetch_all_active_overrides(month_key, user_id)
    accounts = fetch_all_accounts(user_id)
    account_map = {a["id"]: a for a in accounts}
    debts = fetch_all_debts(user_id)
    debt_map = {d["id"]: d for d in debts}

    # Always load prior-month entries so we can show per-item inheritance
    prior_entries = fetch_prior_month_budget_entries(month_key, user_id)
    prior_entry_map = {e["budget_item_id"]: e for e in prior_entries}

    _income_sec = next((s for s in db_sections if "income" in s["key"].lower()), None)
    income_key = _income_sec["key"] if _income_sec else (db_sections[0]["key"] if db_sections else "income")

    today_key = _default_month_key()
    sections = []
    section_totals = {}

    for sec in db_sections:
        section_key = sec["key"]
        section_items = []
        for item in items:
            if item["section"] != section_key:
                continue

            # For debt-linked items: hide in future months past the payoff date
            linked_debt = debt_map.get(item["linked_debt_id"]) if item["linked_debt_id"] else None
            if item["linked_debt_id"] and month_key > today_key:
                payoff_mk = _debt_payoff_month_key(linked_debt) if linked_debt else None
                if payoff_mk and month_key > payoff_mk:
                    continue

            linked_account = account_map.get(item["linked_account_id"]) if item["linked_account_id"] else None
            is_linked_account = linked_account is not None
            is_linked_debt = linked_debt is not None
            override_reason = ""

            if item["id"] in entry_map:
                amount = float(entry_map[item["id"]]["amount"] or 0)
                source = "manual_override"
            elif (
                is_linked_account
                and linked_account["id"] in active_overrides
            ):
                amount = float(active_overrides[linked_account["id"]]["override_amount"] or 0)
                override_reason = (active_overrides[linked_account["id"]]["reason"] or "").strip()
                source = "manual_override"
            elif is_linked_account:
                amount = float(linked_account["monthly_contribution"] or 0)
                source = "linked_account"
            elif is_linked_debt:
                amount = float(linked_debt.get("monthly_payment") or 0)
                source = "linked_debt"
            elif item["id"] in prior_entry_map:
                amount = float(prior_entry_map[item["id"]]["amount"] or 0)
                source = "inherited"
            else:
                amount = float(item["default_amount"] or 0)
                source = "default"

            if source == "default":
                source_label = "default"
                source_title = "Using the default amount — change it to save for this month"
            elif source == "inherited":
                source_label = "prev month"
                source_title = "Carried forward from the last month you edited"
            elif source == "manual_override":
                source_label = "this month"
                source_title = "Saved for this month"
                if is_linked_account:
                    ln = (linked_account.get("name") or "linked account").strip()
                    if override_reason and override_reason != "from budget":
                        source_label = "Payment calendar"
                        source_title = f"Pulled from your Payment calendar for this month ({override_reason})"
                    else:
                        source_title = f"Saved for this month (linked account · {ln})"
                elif is_linked_debt:
                    source_title = "Saved for this month (linked debt)"
            elif source == "linked_account":
                ln = (linked_account.get("name") or "—").strip()
                source_label = f"linked account · {ln}"
                source_title = "Pulled from your linked account's monthly payment"
            else:  # linked_debt
                source_label = "linked debt"
                source_title = "Pulled from your Debts page — update the monthly payment there to change this amount"

            linked_account_name = None
            pre_salary = False
            if linked_account:
                linked_account_name = linked_account["name"]
                pre_salary = bool(linked_account.get("pre_salary"))

            section_items.append({
                "id": item["id"],
                "name": item["name"],
                "notes": item["notes"],
                "amount": amount,
                "annual_amount": amount * 12,
                "default_amount": float(item["default_amount"] or 0),
                "section": section_key,
                "linked_account": item["linked_account_id"] is not None,
                "linked_account_id": item["linked_account_id"],
                "linked_account_name": linked_account_name,
                "linked_debt": linked_debt is not None,
                "linked_debt_id": item["linked_debt_id"],
                "linked_debt_name": linked_debt["name"] if linked_debt else None,
                "source": source,
                "source_label": source_label,
                "source_title": source_title,
                "pre_salary": pre_salary,
            })
        section_total = sum(i["amount"] for i in section_items)
        section_totals[section_key] = section_total
        sections.append({
            "key": section_key,
            "label": sec["label"],
            "rows": section_items,
            "total": section_total,
        })

    # Pre-salary items (salary sacrifice) are shown for visibility but never flowed
    # through take-home pay, so exclude them from the surplus deduction.
    pre_salary_total = sum(
        item["amount"]
        for sec in sections
        for item in sec["rows"]
        if item.get("pre_salary")
    )
    total_income = section_totals.get(income_key, 0)
    for section in sections:
        section["income_percent"] = (section["total"] / total_income * 100) if total_income > 0 else None
        section["annual_total"] = section["total"] * 12
    total_expenses = sum(v for k, v in section_totals.items() if k != income_key)
    surplus = total_income - (total_expenses - pre_salary_total)
    savings_total = 0.0
    for sec in db_sections:
        k = sec["key"]
        if k != income_key and ("invest" in k or "saving" in k):
            savings_total += section_totals.get(k, 0)
    savings_rate = (savings_total / total_income * 100) if total_income > 0 else 0.0

    summary = {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "pre_salary_total": pre_salary_total,
        "surplus": surplus,
        "savings_rate": savings_rate,
    }

    return sections, summary


@budget_bp.route("/", methods=["GET", "POST"])
@login_required
def budget():
    uid = current_user.id
    month_key = valid_month_key(request.values.get("month")) or _default_month_key()

    if request.method == "POST":
        item_id = request.form.get("item_id", type=int)
        if item_id:
            amount = optional_float(request.form.get("amount"), 0.0)
            upsert_budget_entry(month_key, item_id, amount, uid)
            _sync_linked_override(item_id, month_key, amount, uid)
        return redirect(url_for("budget.budget", month=month_key))

    sections, summary = _build_monthly_data(month_key, uid)
    db_sections = fetch_budget_sections(uid)
    all_items = fetch_budget_items(uid)
    has_budget_basics = any(
        (not _row_value(item, "linked_account_id")) or float(_row_value(item, "default_amount", 0) or 0) > 0
        for item in all_items
    )
    budget_setup_href = (
        url_for("budget.budget_items_view", mode="create", focus="first_budget", month=month_key)
        if not has_budget_basics
        else url_for("budget.budget_items_view", month=month_key)
    )
    _income_sec = next((s for s in db_sections if "income" in s["key"].lower()), None)
    income_key = _income_sec["key"] if _income_sec else (db_sections[0]["key"] if db_sections else "income")

    month_strip, is_inherited = _build_budget_month_strip(month_key, uid)

    return render_template(
        "budget.html",
        month_key=month_key,
        month_label=_month_label(month_key),
        monthly_update_href=url_for("monthly_review.monthly_review", month=month_key),
        sections=sections,
        summary=summary,
        income_key=income_key,
        month_strip=month_strip,
        is_inherited=is_inherited,
        budget_setup_href=budget_setup_href,
        has_budget_basics=has_budget_basics,
        active_page="budget",
    )


@budget_bp.route("/contribution-calendar", methods=["GET", "POST"])
@login_required
def contribution_calendar():
    uid = current_user.id
    default_from_month, default_to_month = _default_contribution_calendar_range()
    selected_from_month = valid_month_key(request.values.get("from_month")) or default_from_month
    raw_to_month = valid_month_key(request.values.get("to_month"))
    if raw_to_month:
        selected_to_month = raw_to_month
    elif selected_from_month == default_from_month:
        selected_to_month = default_to_month
    else:
        selected_to_month = add_months_to_key(selected_from_month, 23)
    if selected_to_month < selected_from_month:
        selected_to_month = add_months_to_key(selected_from_month, 23)

    redirect_args = {
        "from_month": selected_from_month,
        "to_month": selected_to_month,
    }

    if request.method == "POST":
        form_name = (request.form.get("form_name") or "").strip()

        if form_name == "create_temporary_plan":
            plan_name = (request.form.get("plan_name") or "").strip()
            start_month = valid_month_key(request.form.get("start_month"))
            end_month = valid_month_key(request.form.get("end_month"))
            invalid_amount_fields = []
            rows = []

            for key, raw_value in request.form.items():
                if not key.startswith("amount_"):
                    continue
                entry_key = key[len("amount_"):]
                account_id, component = _parse_calendar_entry_key(entry_key)
                if account_id is None:
                    continue
                amount_text = (raw_value or "").strip()
                if not amount_text:
                    continue
                amount = optional_float(amount_text, default=None, min_val=0.0)
                if amount is None:
                    invalid_amount_fields.append(entry_key)
                    continue
                rows.append({
                    "account_id": account_id,
                    "component": component,
                    "from_month": start_month,
                    "to_month": end_month,
                    "override_amount": amount,
                })

            if not plan_name:
                flash("Name the temporary plan so it can be grouped and removed later.", "error")
            elif not start_month or not end_month:
                flash("Choose a valid start and end month for the temporary plan.", "error")
            elif end_month < start_month:
                flash("The end month must be the same as or after the start month.", "error")
            elif invalid_amount_fields:
                flash("One or more plan amounts were not valid numbers.", "error")
            elif not rows:
                flash("Add at least one account amount to create a temporary plan.", "error")
            else:
                created = create_temporary_contribution_plan(uid, plan_name, rows)
                if created.get("created_count"):
                    flash(
                        f"Saved temporary plan '{created['plan_name']}' for {created['created_count']} account override"
                        f"{'s' if created['created_count'] != 1 else ''}.",
                        "success",
                    )
                else:
                    flash("No eligible account rows were saved for that temporary plan.", "error")

            return redirect(url_for("budget.contribution_calendar", **redirect_args))

        if form_name == "create_annual_pot_fill_plan":
            plan_name = (request.form.get("plan_name") or "").strip()
            account_key = request.form.get("pattern_account_key") or request.form.get("pattern_account_id")
            account_id, component = _parse_calendar_entry_key(account_key)
            start_month = valid_month_key(request.form.get("pattern_start_month"))
            months_per_year = optional_int(request.form.get("pattern_months_per_year"), 0) or 0
            years = optional_int(request.form.get("pattern_years"), 0) or 0
            monthly_amount = optional_float(request.form.get("pattern_monthly_amount"), default=None, min_val=0.0)
            amount_sequence = []
            invalid_sequence = False
            sequence_text = (request.form.get("pattern_monthly_amounts") or "").strip()
            if sequence_text:
                for part in sequence_text.split(","):
                    amount = optional_float(part.strip(), default=None, min_val=0.0)
                    if amount is None:
                        invalid_sequence = True
                        break
                    amount_sequence.append(amount)
                months_per_year = len(amount_sequence)
            rows = []
            if account_id and start_month and months_per_year > 0 and years > 0 and not invalid_sequence:
                months_per_year = min(months_per_year, 12)
                years = min(years, 10)
                for year_idx in range(years):
                    year_start_month = add_months_to_key(start_month, year_idx * 12)
                    if amount_sequence:
                        for month_idx, amount in enumerate(amount_sequence[:months_per_year]):
                            month_key = add_months_to_key(year_start_month, month_idx)
                            rows.append({
                                "account_id": account_id,
                                "component": component,
                                "from_month": month_key,
                                "to_month": month_key,
                                "override_amount": amount,
                            })
                    elif monthly_amount is not None:
                        rows.append({
                            "account_id": account_id,
                            "component": component,
                            "from_month": year_start_month,
                            "to_month": add_months_to_key(year_start_month, months_per_year - 1),
                            "override_amount": monthly_amount,
                        })

            if not plan_name:
                flash("Name the yearly pot-fill plan so it can be grouped and removed later.", "error")
            elif not account_id:
                flash("Choose which account the yearly pot-fill plan belongs to.", "error")
            elif not start_month:
                flash("Choose a valid first start month for the yearly pot-fill plan.", "error")
            elif invalid_sequence:
                flash("Use only valid comma-separated amounts for the optional month-by-month pattern.", "error")
            elif months_per_year <= 0 or years <= 0 or (monthly_amount is None and not amount_sequence):
                flash("Enter a valid monthly amount, months per year and number of years.", "error")
            else:
                created = create_temporary_contribution_plan(uid, plan_name, rows)
                if created.get("created_count"):
                    flash(
                        f"Saved yearly pot-fill plan '{created['plan_name']}' for {created['created_count']} yearly block"
                        f"{'s' if created['created_count'] != 1 else ''}.",
                        "success",
                    )
                else:
                    flash("No eligible yearly pot-fill rows were saved.", "error")

            return redirect(url_for("budget.contribution_calendar", **redirect_args))

        if form_name == "delete_temporary_plan":
            reason = (request.form.get("reason") or request.form.get("plan_name") or "").strip()
            deleted = delete_temporary_contribution_plan(uid, reason)
            if deleted:
                flash(f"Removed {deleted} temporary override row{'s' if deleted != 1 else ''}.", "success")
            else:
                flash("That temporary plan was not found.", "error")
            return redirect(url_for("budget.contribution_calendar", **redirect_args))

        return redirect(url_for("budget.contribution_calendar", **redirect_args))

    calendar = fetch_contribution_calendar(uid, selected_from_month, selected_to_month)
    plans = fetch_temporary_contribution_plans(uid)
    assumptions = fetch_assumptions(uid)
    pension_carry_forward_entries = fetch_pension_carry_forward(uid)
    allowance_frame = _build_contribution_allowance_frame(calendar, assumptions, pension_carry_forward_entries)
    month_columns = [
        {"key": month_key, "label": datetime.strptime(month_key, "%Y-%m").strftime("%b %Y")}
        for month_key in calendar["months"]
    ]

    return render_template(
        "budget_contribution_calendar.html",
        from_month=selected_from_month,
        to_month=selected_to_month,
        month_columns=month_columns,
        calendar=calendar,
        allowance_frame=allowance_frame,
        plans=plans,
        active_page="budget",
        monthly_update_href=url_for("monthly_review.monthly_review", month=selected_from_month),
    )


def _stamp_inherited_entries(month_key, user_id):
    """On first save to a new month, copy prior-month non-linked entries so the
    month owns its own values and isn't silently affected by later edits to the
    prior month.
    """
    existing = fetch_budget_entries(month_key, user_id)
    if existing:
        return
    prior_entries = fetch_prior_month_budget_entries(month_key, user_id)
    if not prior_entries:
        return
    items = fetch_budget_items(user_id)
    linked_ids = {it["id"] for it in items if it["linked_account_id"] or it["linked_debt_id"]}
    for entry in prior_entries:
        if entry["budget_item_id"] not in linked_ids:
            upsert_budget_entry(month_key, entry["budget_item_id"], float(entry["amount"] or 0), user_id)


@budget_bp.route("/api/entry", methods=["POST"])
@login_required
def budget_save_entry():
    """AJAX endpoint — save a single budget entry, return JSON."""
    uid = current_user.id
    month_key = valid_month_key(request.form.get("month")) or _default_month_key()
    item_id = request.form.get("item_id", type=int)
    amount = optional_float(request.form.get("amount"), 0.0)
    if item_id:
        _stamp_inherited_entries(month_key, uid)
        upsert_budget_entry(month_key, item_id, amount, uid)
        _sync_linked_override(item_id, month_key, amount, uid)
    return jsonify({"ok": True})


@budget_bp.route("/api/contribution-override", methods=["POST"])
@login_required
def budget_save_contribution_override():
    """AJAX endpoint — save or delete a single month contribution override, return JSON."""
    uid = current_user.id
    month_key = valid_month_key(request.form.get("month"))
    account_id = request.form.get("account_id", type=int)
    
    amount_raw = request.form.get("amount")
    
    if month_key and account_id:
        if amount_raw is None or str(amount_raw).strip() == "":
            from app.models.planning_allowances import delete_single_month_contribution_override
            delete_single_month_contribution_override(account_id, month_key, uid)
        else:
            amount = optional_float(amount_raw, 0.0)
            upsert_single_month_contribution_override(
                account_id, month_key, amount, uid, reason="Calendar edit"
            )
    return jsonify({"ok": True})


@budget_bp.route("/api/quick-add", methods=["POST"])
@login_required
def budget_quick_add():
    """AJAX endpoint — add a new budget item from the budget view."""
    uid = current_user.id
    name = (request.form.get("name") or "").strip()
    section = (request.form.get("section") or "").strip()
    if not name or not section:
        return jsonify({"ok": False, "error": "Name and section required"}), 400

    existing = fetch_budget_items(uid)
    sort_order = max(
        (i["sort_order"] for i in existing if i["section"] == section), default=-1
    ) + 1
    item_id = create_budget_item({
        "name": name,
        "section": section,
        "default_amount": 0.0,
        "linked_account_id": None,
        "notes": "",
        "sort_order": sort_order,
    }, uid)
    return jsonify({"ok": True, "item_id": item_id, "name": name})


@budget_bp.route("/import", methods=["POST"])
@login_required
def budget_import():
    """Import budget items and amounts from an uploaded .xlsx file.

    Reads the SteadyPlan export format:
      Row 1: title
      Row 2: generated date
      Then repeating blocks of:
        - Section header row (col A = section label, col C = "Amount")
        - Item rows (col A = name, col B = notes, col C = amount)
        - "Section total" row
        - Blank rows
      Finally summary rows (Total Income, Total Expenses, Surplus)
    """
    uid = current_user.id
    month_key = valid_month_key(request.form.get("month")) or _default_month_key()

    f = request.files.get("file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        flash("Please upload an .xlsx file.", "error")
        return redirect(url_for("budget.budget", month=month_key))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        current_app.logger.warning("budget import: could not read workbook: %s", e)
        flash("Could not read the Excel file.", "error")
        return redirect(url_for("budget.budget", month=month_key))

    # Load existing sections and items for matching
    db_sections = fetch_budget_sections(uid)
    section_label_to_key = {s["label"].strip().lower(): s["key"] for s in db_sections}
    existing_items = fetch_budget_items(uid)

    current_section_key = None
    items_imported = 0
    sections_created = 0
    skip_labels = {"total income", "total expenses", "surplus", "section total", ""}

    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = [cell.value for cell in row]
        col_a = str(vals[0] or "").strip()
        col_b = str(vals[1] or "").strip()
        col_c = vals[2] if len(vals) > 2 else None

        # Skip title, date, and empty rows
        if not col_a and not col_c:
            continue

        # Detect section header: col C is literally "Amount"
        if isinstance(col_c, str) and col_c.strip().lower() == "amount":
            label = col_a
            label_lower = label.lower()
            if label_lower in section_label_to_key:
                current_section_key = section_label_to_key[label_lower]
            else:
                # Create a new section
                new_key = create_budget_section(label, uid)
                section_label_to_key[label_lower] = new_key
                current_section_key = new_key
                sections_created += 1
                # Refresh sections
                db_sections = fetch_budget_sections(uid)
            continue

        # Skip summary/total rows
        if col_a.lower() in skip_labels or col_b.lower() == "section total":
            continue

        # Skip rows without a numeric amount
        amount = None
        if isinstance(col_c, (int, float)):
            amount = float(col_c)
        else:
            try:
                cleaned = str(col_c or "").replace("£", "").replace(",", "").strip()
                amount = float(cleaned) if cleaned else None
            except (ValueError, TypeError):
                amount = None

        if amount is None or not current_section_key or not col_a:
            continue

        # Find or create the budget item
        item_name = col_a
        notes = col_b if col_b else ""
        matched_item = None
        for it in existing_items:
            if it["name"].strip().lower() == item_name.lower() and it["section"] == current_section_key:
                matched_item = it
                break

        if not matched_item:
            # Create a new item
            sort_order = max(
                (i["sort_order"] for i in existing_items if i["section"] == current_section_key),
                default=-1,
            ) + 1
            item_id = create_budget_item({
                "name": item_name,
                "section": current_section_key,
                "default_amount": amount,
                "linked_account_id": None,
                "notes": notes,
                "sort_order": sort_order,
            }, uid)
            # Refresh items list
            existing_items = fetch_budget_items(uid)
        else:
            item_id = matched_item["id"]

        # Upsert the entry for this month
        upsert_budget_entry(month_key, item_id, amount, uid)
        items_imported += 1

    msg = f"Imported {items_imported} budget items"
    if sections_created:
        msg += f" and created {sections_created} new section{'s' if sections_created > 1 else ''}"
    msg += f" for {_month_label(month_key)}."
    flash(msg, "success")
    return redirect(url_for("budget.budget", month=month_key))


# ── Annual import (12-month workbook produced by annual-export) ──────────────

_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _sheet_name_to_month_key(sheet_name):
    """Parse 'Apr 2026' → '2026-04'. Returns None if the name doesn't match."""
    parts = (sheet_name or "").strip().split()
    if len(parts) != 2:
        return None
    name, yr = parts
    if name not in _MONTH_NAMES:
        return None
    try:
        year = int(yr)
    except ValueError:
        return None
    month = _MONTH_NAMES.index(name) + 1
    return f"{year:04d}-{month:02d}"


def _read_annual_month_sheet(ws, existing_items_by_id, existing_items_by_name):
    """Read rows off an annual-export month tab (cols: item_id, name, notes, amount).

    Returns a list of (matched_item, amount) tuples for rows we understood,
    plus a list of unmatched (name, amount) tuples for the preview so the user
    sees what was skipped and why.
    """
    matched = []
    unmatched = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4:
            continue
        id_cell, name_cell, _notes_cell, amount_cell = row[0], row[1], row[2], row[3]

        # Skip section headers (col D = literal "Amount") and section-total rows.
        if isinstance(amount_cell, str):
            continue
        # Skip fully-blank rows between sections.
        if amount_cell is None and not isinstance(name_cell, str):
            continue

        matched_item = None
        if isinstance(id_cell, int) and id_cell in existing_items_by_id:
            matched_item = existing_items_by_id[id_cell]
        elif isinstance(name_cell, str):
            matched_item = existing_items_by_name.get(name_cell.strip().lower())

        try:
            amount = float(amount_cell) if amount_cell is not None else 0.0
        except (TypeError, ValueError):
            continue

        if matched_item:
            matched.append((matched_item, amount))
        elif isinstance(name_cell, str) and name_cell.strip():
            unmatched.append((name_cell.strip(), amount))

    return matched, unmatched


def _compute_annual_import_diff(wb, uid):
    """Walk the workbook's month tabs, compare proposed amounts against current
    DB state, and return a structured diff suitable for a preview screen and a
    later apply pass.

    Output shape:
        {
          "months": [
            {
              "month_key": "2026-07",
              "month_label": "Jul 2026",
              "changes": [{"item_id", "item_name", "section", "old", "new", "linked"}, ...],
              "unchanged_count": int,
            },
            ...
          ],
          "unknown_items_by_month": {"2026-07": ["Foo", "Bar"], ...},
          "unrecognised_sheets": ["Sheet1", ...],
        }
    """
    existing_items = fetch_budget_items(uid)
    items_by_id = {it["id"]: it for it in existing_items}
    items_by_name = {it["name"].strip().lower(): it for it in existing_items}

    months = []
    unknown_items_by_month = {}
    unrecognised_sheets = []

    for sheet_name in wb.sheetnames:
        if sheet_name in ("Summary", "Investment Tracking"):
            continue
        month_key = _sheet_name_to_month_key(sheet_name)
        if not month_key:
            unrecognised_sheets.append(sheet_name)
            continue

        ws = wb[sheet_name]
        matched, unmatched = _read_annual_month_sheet(ws, items_by_id, items_by_name)

        # Current DB state for this month, to build the diff
        current_entries = fetch_budget_entries(month_key, uid)
        current_by_item = {e["budget_item_id"]: float(e["amount"] or 0) for e in current_entries}

        changes = []
        unchanged_count = 0
        for item, new_amount in matched:
            old_amount = current_by_item.get(item["id"])
            # If there's no current entry but the workbook amount equals the item's
            # default, treat that as unchanged too — the round-trip would otherwise
            # look noisy with every untouched carry-forward row showing up.
            effective_old = old_amount if old_amount is not None else float(item["default_amount"] or 0)
            if abs(effective_old - new_amount) < 0.005:
                unchanged_count += 1
                continue
            changes.append({
                "item_id": item["id"],
                "item_name": item["name"],
                "section": item["section"],
                "old": effective_old,
                "old_is_entry": old_amount is not None,
                "new": new_amount,
                "linked": bool(item["linked_account_id"]),
            })

        if changes or unmatched:
            months.append({
                "month_key": month_key,
                "month_label": _month_label(month_key),
                "changes": changes,
                "unchanged_count": unchanged_count,
            })
        if unmatched:
            unknown_items_by_month[month_key] = [name for name, _ in unmatched]

    return {
        "months": months,
        "unknown_items_by_month": unknown_items_by_month,
        "unrecognised_sheets": unrecognised_sheets,
    }


@budget_bp.route("/annual-import", methods=["POST"])
@login_required
def budget_annual_import():
    """Stage 1 of annual upload: parse the workbook, compute a diff against current
    budget state, and show a preview. No DB writes yet — user must confirm via the
    /annual-import/confirm route before anything lands.
    """
    uid = current_user.id

    f = request.files.get("file")
    if not f or not f.filename.endswith((".xlsx", ".xls")):
        flash("Please upload an .xlsx file.", "error")
        return redirect(url_for("budget.budget"))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(f.read()), data_only=True)
    except Exception as e:
        current_app.logger.warning("annual-import: could not read workbook: %s", e)
        flash("Could not read the Excel file.", "error")
        return redirect(url_for("budget.budget"))

    diff = _compute_annual_import_diff(wb, uid)

    total_changes = sum(len(m["changes"]) for m in diff["months"])

    # Any previously-staged diff for this user is now stale — drop it whether
    # or not we proceed.
    _drop_staged_annual_import(uid)

    if total_changes == 0:
        flash(
            "No changes detected — the workbook matches your current budget. Nothing to import.",
            "info",
        )
        return redirect(url_for("budget.budget"))

    # Persist the minimal replay payload to disk (survives the 4KB cookie
    # limit). The session only carries the token + owning user_id so a
    # tampered cookie can't confirm someone else's staging.
    payload = {
        "user_id": uid,
        "changes": [
            {"item_id": c["item_id"], "month_key": m["month_key"],
             "new": c["new"], "linked": c["linked"]}
            for m in diff["months"]
            for c in m["changes"]
        ],
    }
    token = write_staged(current_app, payload)
    session["budget_annual_import"] = {"token": token, "user_id": uid}

    return render_template(
        "budget_annual_import_preview.html",
        diff=diff,
        total_changes=total_changes,
        active_page="budget",
    )


def _load_staged_annual_import(uid):
    """Load the staged diff for `uid` or return None if missing/stale/mismatched."""
    stash = session.get("budget_annual_import")
    if not stash or stash.get("user_id") != uid:
        return None
    token = stash.get("token")
    payload = read_staged(current_app, token) if token else None
    if not payload or payload.get("user_id") != uid:
        return None
    return payload


def _drop_staged_annual_import(uid):
    """Best-effort cleanup of any staged diff belonging to `uid`."""
    stash = session.pop("budget_annual_import", None)
    if stash and stash.get("token"):
        delete_staged(current_app, stash["token"])


@budget_bp.route("/annual-import/confirm", methods=["POST"])
@login_required
def budget_annual_import_confirm():
    """Stage 2: apply the previously-staged diff. Re-validates ownership at each
    write via upsert_budget_entry + _sync_linked_override."""
    uid = current_user.id
    payload = _load_staged_annual_import(uid)
    if not payload or not payload.get("changes"):
        _drop_staged_annual_import(uid)
        flash("Nothing to import — the preview expired. Upload the file again.", "info")
        return redirect(url_for("budget.budget"))

    total_written = 0
    months_touched = set()
    linked_syncs = []
    for change in payload["changes"]:
        try:
            item_id = int(change["item_id"])
            amount = float(change["new"])
        except (TypeError, ValueError, KeyError):
            continue
        month_key = change.get("month_key")
        if not valid_month_key(month_key):
            continue
        upsert_budget_entry(month_key, item_id, amount, uid)
        total_written += 1
        months_touched.add(month_key)
        if change.get("linked"):
            linked_syncs.append((item_id, month_key, amount))

    for item_id, month_key, amount in linked_syncs:
        _sync_linked_override(item_id, month_key, amount, uid)

    _drop_staged_annual_import(uid)
    flash(
        f"Imported {total_written} entries across {len(months_touched)} month"
        f"{'s' if len(months_touched) != 1 else ''}.",
        "success",
    )
    return redirect(url_for("budget.budget"))


@budget_bp.route("/annual-import/cancel", methods=["POST"])
@login_required
def budget_annual_import_cancel():
    """Discard a staged annual-import diff."""
    _drop_staged_annual_import(current_user.id)
    flash("Annual import cancelled — nothing was changed.", "info")
    return redirect(url_for("budget.budget"))


@budget_bp.route("/trend/")
@login_required
def budget_trend():
    uid = current_user.id
    today = date.today()

    # Last 6 months that have any entries
    all_months = _last_n_months(today, 6)
    months_with_data = fetch_months_with_budget_entries(uid)
    months = [m for m in all_months if m in months_with_data]

    if not months:
        return render_template(
            "budget_trend.html",
            sections={},
            months=[],
            month_labels=[],
            current_month_num=today.month,
            trend_avg_income=0,
            trend_avg_spend=0,
            trend_surplus=0,
            active_page="budget",
        )

    # Load sections for ordering and income detection
    db_sections = fetch_budget_sections(uid)
    section_order_map = {s["label"]: s["sort_order"] for s in db_sections}
    income_section = next((s for s in db_sections if "income" in s["key"].lower()), None)
    income_section_label = income_section["label"] if income_section else None

    # Only use actual recorded entries — no fallback to current defaults
    raw = fetch_budget_trend(uid, months)

    # Build sections: {section_name: {item_name: {month_key: amount, "avg": float}}}
    sections = {}
    for row in raw:
        sn  = row["section_name"]
        inn = row["item_name"]
        mk  = row["month_key"]
        amt = float(row["actual_amount"] or 0)
        if sn not in sections:
            sections[sn] = {}
        if inn not in sections[sn]:
            sections[sn][inn] = {}
        sections[sn][inn][mk] = amt

    # Sort sections by sort_order
    sections = dict(sorted(sections.items(), key=lambda x: section_order_map.get(x[0], 99)))

    # Averages (only over months that have an actual entry)
    for sn, items in sections.items():
        for inn, data in items.items():
            month_vals = [data[mk] for mk in months if mk in data]
            data["avg"] = sum(month_vals) / len(month_vals) if month_vals else 0

    month_labels = [
        datetime.strptime(mk, "%Y-%m").strftime("%b %Y") for mk in months
    ]

    # Hero stats
    trend_avg_income = 0.0
    trend_avg_spend = 0.0
    trend_item_count = 0
    for sn, items in sections.items():
        is_income = (income_section_label and sn == income_section_label)
        for inn, data in items.items():
            trend_item_count += 1
            if is_income:
                trend_avg_income += data["avg"]
            else:
                trend_avg_spend += data["avg"]

    return render_template(
        "budget_trend.html",
        sections=sections,
        months=months,
        month_labels=month_labels,
        trend_avg_income=trend_avg_income,
        trend_avg_spend=trend_avg_spend,
        trend_surplus=trend_avg_income - trend_avg_spend,
        trend_item_count=trend_item_count,
        active_page="budget",
    )


def _build_budget_month_strip(month_key, uid):
    """Build the 12-month tax-year strip (Apr→Mar) for the budget views."""
    saved_months = fetch_months_with_budget_entries(uid)
    mk_year, mk_month = int(month_key[:4]), int(month_key[5:7])
    ty_start_year = mk_year if mk_month >= 4 else mk_year - 1
    today_key = _default_month_key()

    month_strip = []
    for i in range(12):
        m = 4 + i
        y = ty_start_year if m <= 12 else ty_start_year + 1
        if m > 12:
            m -= 12
        mk = f"{y}-{m:02d}"
        month_strip.append({
            "key": mk,
            "label": datetime.strptime(mk, "%Y-%m").strftime("%b"),
            "has_data": mk in saved_months,
            "is_current": mk == month_key,
            "is_today": mk == today_key,
            "month_num": m,
        })

    return month_strip, month_key not in saved_months  # (strip, is_inherited)


def _last_n_months(today, n):
    """Return list of 'YYYY-MM' strings for the last n months (most recent last)."""
    result = []
    y, m = today.year, today.month
    for _ in range(n):
        result.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    result.reverse()
    return result


@budget_bp.route("/items/", methods=["GET", "POST"])
@login_required
def budget_items_view():
    uid = current_user.id
    if request.method == "POST":
        form_name = request.form.get("form_name", "")
        month_key = valid_month_key(request.form.get("month")) or _default_month_key()

        if form_name == "clear_section":
            delete_budget_items_by_section(request.form.get("section_key", ""), uid)
            return redirect(url_for("budget.budget_items_view", month=month_key))

        if form_name == "add_section":
            label = request.form.get("section_label", "").strip()
            if label:
                create_budget_section(label, uid)
            return redirect(url_for("budget.budget_items_view", month=month_key))

        if form_name == "edit_section":
            update_budget_section(
                request.form.get("section_key", ""),
                request.form.get("section_label", "").strip(),
                uid,
            )
            return redirect(url_for("budget.budget_items_view", month=month_key))

        if form_name == "delete_section":
            delete_budget_section(request.form.get("section_key", ""), uid)
            return redirect(url_for("budget.budget_items_view", month=month_key))

        # default: create item
        section = request.form.get("section", "")
        existing = fetch_budget_items(uid)
        sort_order = max(
            (i["sort_order"] for i in existing if i["section"] == section), default=-1
        ) + 1
        create_budget_item({
            "name": request.form.get("name", "").strip(),
            "section": section,
            "default_amount": optional_float(request.form.get("default_amount"), 0.0),
            "linked_account_id": None,
            "notes": request.form.get("notes", "").strip(),
            "sort_order": sort_order,
        }, uid)
        return redirect(url_for("budget.budget_items_view", month=month_key))

    db_sections = fetch_budget_sections(uid)
    all_items = fetch_budget_items(uid)
    accounts = fetch_all_accounts(uid)
    account_map = {a["id"]: dict(a) for a in accounts}

    grouped = []
    for sec in db_sections:
        section_key = sec["key"]
        section_items = []
        for item in all_items:
            if item["section"] != section_key:
                continue
            row = dict(item)
            row["linked_account_name"] = (
                account_map[item["linked_account_id"]]["name"]
                if item["linked_account_id"] and item["linked_account_id"] in account_map
                else None
            )
            section_items.append(row)
        grouped.append({"key": section_key, "label": sec["label"], "rows": section_items})

    selected_id = request.args.get("item_id", type=int)
    selected = fetch_budget_item(selected_id, uid) if selected_id else None
    page_mode = request.args.get("mode", "view" if selected_id else "list")
    section_options = [(s["key"], s["label"]) for s in db_sections]
    has_budget_basics = any(
        (not _row_value(item, "linked_account_id")) or float(_row_value(item, "default_amount", 0) or 0) > 0
        for item in all_items
    )
    first_budget_focus = request.args.get("focus") == "first_budget" and not has_budget_basics
    first_budget_section = next((s["key"] for s in db_sections if "income" in s["key"].lower()), None)

    # Track the month the user came from so "Back to Budget" returns there.
    month_key = valid_month_key(request.args.get("month")) or _default_month_key()
    budget_create_href = (
        url_for("budget.budget_items_view", mode="create", focus="first_budget", month=month_key)
        if not has_budget_basics
        else url_for("budget.budget_items_view", mode="create", month=month_key)
    )

    return render_template(
        "budget_items.html",
        grouped=grouped,
        accounts=accounts,
        section_options=section_options,
        selected=selected,
        page_mode=page_mode,
        active_page="budget",
        month_key=month_key,
        month_label=_month_label(month_key),
        first_budget_focus=first_budget_focus,
        first_budget_section=first_budget_section,
        has_budget_basics=has_budget_basics,
        budget_create_href=budget_create_href,
    )


@budget_bp.route("/items/<int:item_id>", methods=["POST"])
@login_required
def budget_item_action(item_id):
    uid = current_user.id
    month_key = valid_month_key(request.form.get("month")) or _default_month_key()
    if request.form.get("form_name") == "delete":
        delete_budget_item(item_id, uid)
        return redirect(url_for("budget.budget_items_view", month=month_key))

    item = fetch_budget_item(item_id, uid)
    if not item:
        flash("Budget item not found.", "error")
        return redirect(url_for("budget.budget_items_view", month=month_key))

    old_default = float(item["default_amount"] or 0)
    new_default = max(0.0, optional_float(request.form.get("default_amount"), 0.0))

    ok = update_budget_item({
        "id": item_id,
        "name": request.form.get("name", "").strip(),
        "section": request.form.get("section", ""),
        "default_amount": new_default,
        "linked_account_id": item["linked_account_id"],
        "notes": request.form.get("notes", "").strip(),
    }, uid)
    if not ok:
        flash("Budget item not found.", "error")
        return redirect(url_for("budget.budget_items_view", month=month_key))

    # If the default changed, update this month's entry too — but only if it
    # still matches the old default (carry-forward), not if the user had
    # deliberately set a different amount for this month.
    if abs(old_default - new_default) > 0.005 and not item["linked_account_id"] and not item["linked_debt_id"]:
        entries = fetch_budget_entries(month_key, uid)
        for e in entries:
            if e["budget_item_id"] == item_id and abs(float(e["amount"] or 0) - old_default) < 0.005:
                upsert_budget_entry(month_key, item_id, new_default, uid)
                break

    return redirect(url_for("budget.budget_items_view", month=month_key))


# ── Debts ─────────────────────────────────────────────────────────────────────

@budget_bp.route("/debts/", methods=["GET", "POST"])
@login_required
def budget_debts():
    uid = current_user.id

    if request.method == "POST":
        form = request.form
        form_name = form.get("form_name")

        if form_name == "create_debt":
            create_debt({
                "name": form.get("name", "").strip(),
                "original_amount": optional_float(form.get("original_amount"), 0.0),
                "current_balance": optional_float(form.get("current_balance"), 0.0),
                "monthly_payment": optional_float(form.get("monthly_payment"), 0.0),
                "apr": optional_float(form.get("apr"), 0.0),
                "notes": form.get("notes", "").strip(),
                "start_date": form.get("start_date", "").strip() or None,
            }, uid)
            return redirect(url_for("budget.budget_debts"))

        if form_name == "update_debt":
            debt_id = optional_int(form.get("debt_id"))
            if debt_id and fetch_debt(debt_id, uid):
                update_debt(debt_id, {
                    "name": form.get("name", "").strip(),
                    "original_amount": optional_float(form.get("original_amount"), 0.0),
                    "current_balance": optional_float(form.get("current_balance"), 0.0),
                    "monthly_payment": optional_float(form.get("monthly_payment"), 0.0),
                    "apr": optional_float(form.get("apr"), 0.0),
                    "notes": form.get("notes", "").strip(),
                    "start_date": form.get("start_date", "").strip() or None,
                }, uid)
            return redirect(url_for("budget.budget_debts"))

        if form_name == "delete_debt":
            debt_id = optional_int(form.get("debt_id"))
            if debt_id and fetch_debt(debt_id, uid):
                delete_debt(debt_id, uid)
            return redirect(url_for("budget.budget_debts"))

        return redirect(url_for("budget.budget_debts"))

    raw_debts = fetch_all_debts(uid)
    debt_cards = [build_debt_card(d) for d in raw_debts]
    guidance_strategy = (request.args.get("strategy") or "avalanche").strip().lower()
    if guidance_strategy not in {"avalanche", "snowball"}:
        guidance_strategy = "avalanche"
    guidance_extra_monthly = max(optional_float(request.args.get("extra_monthly"), 0.0), 0.0)

    payoff_guidance = None
    if debt_cards:
        guidance_comparison = compare_debt_payoff_strategies(debt_cards, guidance_extra_monthly)
        active_guidance = guidance_comparison[guidance_strategy]
        debt_card_map = {debt["id"]: debt for debt in debt_cards}
        ranked_guidance = []
        for idx, step in enumerate(active_guidance["payoff_steps"]):
            debt = debt_card_map.get(step["id"], {})
            if idx == 0:
                reason = "Highest APR first" if guidance_strategy == "avalanche" else "Smallest balance first"
            else:
                reason = "Gets earlier debt payments once those clear"
            ranked_guidance.append({
                "id": step["id"],
                "name": step["name"],
                "current_balance": float(debt.get("current_balance") or 0),
                "apr": float(debt.get("apr") or 0),
                "reason": reason,
                "rolled_monthly_payment": step["rolled_monthly_payment"],
            })

        if guidance_strategy == "avalanche":
            comparison_note = "Highest APR first usually costs less in interest. Smallest balance first can give earlier pay-off milestones."
        else:
            comparison_note = "Smallest balance first can give earlier pay-off milestones. Highest APR first usually costs less in interest."

        payoff_guidance = {
            "strategy": guidance_strategy,
            "extra_monthly": guidance_extra_monthly,
            "ranked_debts": ranked_guidance,
            "total_months": active_guidance["total_months"],
            "total_interest": active_guidance["total_interest"],
            "included_debt_count": active_guidance["included_debt_count"],
            "excluded_debt_count": active_guidance["excluded_debt_count"],
            "excluded_debts": active_guidance["excluded_debts"],
            "comparison_note": comparison_note,
        }

    selected_id = request.args.get("debt_id", type=int)
    page_mode = request.args.get("mode", "view")
    selected_debt_raw = next((d for d in raw_debts if d["id"] == selected_id), None) if selected_id else None
    selected_debt = next((d for d in debt_cards if d["id"] == selected_id), None) if selected_id else None

    # Build amortisation schedule for the selected debt detail view.
    # If auto-tracked (start_date + original_amount set), show the full schedule
    # from loan inception so past payments are visible alongside future ones.
    schedule = []
    payments_made = 0
    total_interest_all = None
    interest_paid = 0

    if selected_debt:
        payments_made = selected_debt.get("payments_made", 0)
        original = selected_debt.get("original_amount", 0)

        if selected_debt["auto_tracked"] and original > 0:
            # Full schedule from original balance — dates anchored at start_date
            schedule = amortisation_schedule(
                original,
                selected_debt["apr"],
                selected_debt["monthly_payment"],
                start_date=schedule_anchor(selected_debt.get("start_date")),
            )
        elif selected_debt["months_remaining"]:
            # Remaining schedule only — dates anchored after payments already made
            schedule = amortisation_schedule(
                selected_debt["current_balance"],
                selected_debt["apr"],
                selected_debt["monthly_payment"],
                start_date=schedule_anchor(
                    selected_debt.get("start_date"),
                    selected_debt.get("payments_made", 0),
                ),
            )

        if schedule:
            total_interest_all = sum(r["interest"] for r in schedule)
            interest_paid = sum(r["interest"] for r in schedule[:payments_made])
        else:
            interest_paid = 0

    return render_template(
        "budget_debts.html",
        debt_cards=debt_cards,
        payoff_guidance=payoff_guidance,
        selected_debt=selected_debt,
        selected_debt_raw=selected_debt_raw,
        schedule=schedule,
        payments_made=payments_made,
        total_interest_all=total_interest_all,
        interest_paid=interest_paid,
        page_mode=page_mode,
        active_page="budget",
    )


@budget_bp.route("/debts/export.xlsx")
@login_required
def budget_debts_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Style constants (match projections export) ──────────────────────────
    TEAL       = "0F766E"
    TEAL_LIGHT = "CCFBF1"
    RED_LIGHT  = "FEE2E2"
    GREEN_LIGHT= "DCFCE7"
    BORDER_CLR = "D1D5DB"

    title_font  = Font(name="Aptos", bold=True, color=TEAL, size=14)
    sub_font    = Font(name="Aptos", color="6B7280", size=10)
    hdr_font    = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
    data_font   = Font(name="Aptos", color="1F2937", size=10)
    bold_font   = Font(name="Aptos", bold=True, color="1F2937", size=10)
    red_font    = Font(name="Aptos", color="DC2626", size=10)
    green_font  = Font(name="Aptos", color="16A34A", size=10)

    hdr_fill    = PatternFill("solid", fgColor=TEAL)
    alt_fill    = PatternFill("solid", fgColor=TEAL_LIGHT)
    red_fill    = PatternFill("solid", fgColor=RED_LIGHT)
    green_fill  = PatternFill("solid", fgColor=GREEN_LIGHT)
    no_fill     = PatternFill(fill_type=None)
    thin_border = Border(bottom=Side(style="thin", color=BORDER_CLR))
    GBP         = '£#,##0.00'
    GBP0        = '£#,##0'

    def col_width(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def hdr_row(ws, row, values, widths=None):
        for i, v in enumerate(values, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[row].height = 22
        if widths:
            for i, w in enumerate(widths, 1):
                col_width(ws, i, w)

    def data_cell(ws, row, col, value, font=None, fill=None, num_fmt=None, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font or (bold_font if bold else data_font)
        c.fill = fill or (alt_fill if row % 2 == 0 else no_fill)
        c.border = thin_border
        c.alignment = Alignment(vertical="center")
        if num_fmt:
            c.number_format = num_fmt
        return c

    uid = current_user.id
    raw_debts = fetch_all_debts(uid)
    debt_cards = [build_debt_card(d) for d in raw_debts]

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"

    c = ws.cell(row=1, column=1, value="SteadyPlan — Debt Tracker")
    c.font = title_font
    ws.merge_cells("A1:H1")

    c = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    c.font = sub_font
    ws.merge_cells("A2:H2")
    ws.row_dimensions[1].height = 28

    hdr_row(ws, 4,
            ["Debt", "Balance", "Monthly Payment", "APR %",
             "Months Left", "Payoff Date", "Total Interest", "Total Cost"],
            widths=[28, 16, 18, 10, 14, 16, 18, 16])

    for i, d in enumerate(debt_cards, 5):
        fill = alt_fill if i % 2 == 0 else no_fill
        data_cell(ws, i, 1, d["name"],        bold=True)
        data_cell(ws, i, 2, d["current_balance"],   num_fmt=GBP)
        data_cell(ws, i, 3, d["monthly_payment"],    num_fmt=GBP)
        data_cell(ws, i, 4, d["apr"],                num_fmt='0.00"%"')
        data_cell(ws, i, 5, d["months_remaining"])
        data_cell(ws, i, 6, d["payoff_date"].strftime("%b %Y") if d["payoff_date"] else "—")
        c = data_cell(ws, i, 7, d["total_interest"],     num_fmt=GBP0)
        c.font = red_font
        data_cell(ws, i, 8, d["total_cost"],         num_fmt=GBP0)

    ws.freeze_panes = "A5"

    # ════════════════════════════════════════════════════════════════════════
    # Per-debt sheets — one tab per scenario
    # ════════════════════════════════════════════════════════════════════════
    for d in debt_cards:
        if not d["months_remaining"]:
            continue

        base_payment = d["monthly_payment"]
        scenarios = [
            ("Base",    0),
            ("+£50",   50),
            ("+£100", 100),
            ("+£200", 200),
            ("Double", base_payment),
        ]

        export_anchor = schedule_anchor(d.get("start_date"), d.get("payments_made", 0))

        for label, extra in scenarios:
            new_payment = base_payment + extra
            sched = amortisation_schedule(d["current_balance"], d["apr"], new_payment, start_date=export_anchor)
            total_interest = sum(r["interest"] for r in sched)
            months_saved = (d["months_remaining"] or 0) - len(sched) if extra > 0 else 0
            interest_saved = (d["total_interest"] or 0) - total_interest if extra > 0 else 0

            # Tab name: e.g. "Car Loan — Base", "Car Loan — +£50"
            tab = f"{d['name'][:20]} — {label}"
            ws2 = wb.create_sheet(title=tab)

            # Title
            title_str = f"{d['name']} — {label}"
            if extra > 0:
                title_str += f" (£{new_payment:,.2f}/mo)"
            c = ws2.cell(row=1, column=1, value=f"SteadyPlan — {title_str}")
            c.font = title_font
            ws2.merge_cells("A1:E1")
            ws2.row_dimensions[1].height = 28

            # Summary box rows 2-5
            summaries = [
                ("Balance",        f"£{d['current_balance']:,.2f}"),
                ("Monthly payment",f"£{new_payment:,.2f}"),
                ("Payoff in",      f"{len(sched)} months"),
                ("Total interest", f"£{total_interest:,.0f}"),
            ]
            if extra > 0:
                summaries += [
                    ("Months saved",   f"{months_saved}"),
                    ("Interest saved", f"£{interest_saved:,.0f}"),
                ]
            for r_i, (k, v) in enumerate(summaries, 2):
                lc = ws2.cell(row=r_i, column=1, value=k)
                lc.font = sub_font
                vc = ws2.cell(row=r_i, column=2, value=v)
                vc.font = bold_font if "saved" not in k else Font(name="Aptos", bold=True, color="16A34A", size=10)

            header_row = len(summaries) + 3
            date_col_header = "Date" if export_anchor else "Month"
            hdr_row(ws2, header_row,
                    [date_col_header, "Payment", "Interest", "To Principal", "Balance"],
                    widths=[14, 16, 16, 16, 16])

            for row_i, row in enumerate(sched, header_row + 1):
                fill = alt_fill if row_i % 2 == 0 else no_fill
                date_label = row["date"].strftime("%-d %b %Y") if "date" in row else row["month"]
                data_cell(ws2, row_i, 1, date_label)
                data_cell(ws2, row_i, 2, row["payment"],   num_fmt=GBP)
                c = data_cell(ws2, row_i, 3, row["interest"],  num_fmt=GBP)
                c.font = red_font
                c = data_cell(ws2, row_i, 4, row["principal"], num_fmt=GBP)
                c.font = green_font
                data_cell(ws2, row_i, 5, row["balance"],   num_fmt=GBP, bold=True)

            ws2.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="debts.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
