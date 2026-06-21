"""
Export routes — generates .xlsx downloads for Projections and Budget.

Clean, professional styling with consistent headers and plain UK-formatted data.
"""
from datetime import date, datetime
from io import BytesIO

from flask import Blueprint, send_file, request
from app.utils import valid_month_key
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Alignment, Side, numbers
from openpyxl.utils import get_column_letter

from app.calculations import (
    _safe_get,
    account_gross_growth_rate,
    account_growth_rate,
    add_months_to_key,
    compute_performance_series,
    contribution_override_for_month,
    contribution_breakdown,
    current_age_from_assumptions,
    effective_account_value,
    effective_fee_pct,
    future_value,
    is_pension_account,
    projected_contribution_breakdown,
    projected_account_value,
    projected_account_value_at_year,
    projected_account_value_at_month,
    projected_account_value_at_month_no_fees,
    projected_account_value_at_year_no_fees,
    projected_account_value_no_fees,
    projection_monthly_contribution,
    projection_start_month_key,
    month_key_to_index,
    projected_total_retirement_value,
    to_float,
    uk_tax_year_end,
    uk_tax_year_label,
    uk_tax_year_start,
    years_to_retirement,
    ISA_WRAPPER_TYPES,
    LISA_WRAPPER_TYPES,
)
from app.models import (
    fetch_all_accounts,
    fetch_all_active_overrides,
    fetch_assumptions,
    fetch_budget_entries,
    fetch_budget_items,
    fetch_budget_sections,
    fetch_contribution_overrides,
    fetch_holding_totals_by_account,
    fetch_isa_contributions,
    fetch_monthly_performance_data,
    fetch_monthly_performance_data_by_account,
    fetch_pension_contributions,
    fetch_prior_month_budget_entries,
)
from app.models.accounts import PREMIUM_BONDS_MAX_BALANCE, is_premium_bonds_account
from app.services.financial_truth import refresh_account_snapshots_for_month
from app.services.planning_insights import classify_account

export_bp = Blueprint("export", __name__)

_PERFORMANCE_EXPORT_PERIODS = {
    "1M": 1,
    "6M": 6,
    "1Y": 12,
    "ALL": None,
}


def _normalise_performance_export_period(period):
    key = str(period or "ALL").upper()
    return key if key in _PERFORMANCE_EXPORT_PERIODS else "ALL"


def _filter_monthly_data_for_period(monthly_data, period):
    filtered = list(monthly_data or [])
    period_key = _normalise_performance_export_period(period)
    months = _PERFORMANCE_EXPORT_PERIODS[period_key]
    if months is None or len(filtered) <= months + 1:
        return period_key, filtered
    return period_key, filtered[-(months + 1):]

# ── Clean Shelly colour palette ──────────────────────────────────────────────
_SHELLY_TEAL   = "0F766E"   # Shelly's signature teal (header bg)
_SHELLY_LIGHT  = "CCFBF1"   # Pale teal tint for alternating rows
_BORDER_COLOUR = "D1D5DB"   # Light grey border

_TITLE_FONT    = Font(name="Aptos", bold=True, color="0F766E", size=14)
_SUBTITLE_FONT = Font(name="Aptos", color="6B7280", size=10)
_HEADER_FONT   = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
_DATA_FONT     = Font(name="Aptos", color="1F2937", size=10)
_DATA_BOLD     = Font(name="Aptos", bold=True, color="1F2937", size=10)
_ACCENT_FONT   = Font(name="Aptos", bold=True, color="0F766E", size=11)

_HEADER_FILL   = PatternFill("solid", fgColor=_SHELLY_TEAL)
_ALT_FILL      = PatternFill("solid", fgColor=_SHELLY_LIGHT)
_NO_FILL       = PatternFill(fill_type=None)

_THIN_BORDER   = Border(
    bottom=Side(style="thin", color=_BORDER_COLOUR),
)


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _display_schedule_to_month(to_month):
    """Render open-ended contribution override sentinels as user-facing text."""
    if not to_month:
        return "Ongoing"
    text = str(to_month)
    return "Ongoing" if text >= "9999-12" else text


def _header_row(ws, row_num, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[row_num].height = 24


def _data_row(ws, row_num, values, bold=False, num_formats=None):
    num_formats = num_formats or {}
    font = _DATA_BOLD if bold else _DATA_FONT
    fill = _ALT_FILL if row_num % 2 == 0 else _NO_FILL
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = _THIN_BORDER
        if col in num_formats:
            cell.number_format = num_formats[col]


def _title_cell(ws, row_num, text, col_span=1):
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = _TITLE_FONT
    if col_span > 1:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_span)


def _tax_year_label_for_month_key(month_key, assumptions=None):
    """Return the UK tax-year label for a projected contribution month.

    SteadyPlan treats April contributions according to the configured salary day:
    if the salary/review day is before 6 April, that April contribution still
    belongs to the previous tax year; otherwise April belongs to the new tax year.
    """
    try:
        year_text, month_text = str(month_key).split("-")
        year = int(year_text)
        month = int(month_text)
    except (AttributeError, TypeError, ValueError):
        return ""

    salary_day = 0
    try:
        salary_day = int(_safe_get(assumptions, "salary_day", 0) or 0)
    except (TypeError, ValueError):
        salary_day = 0

    if month > 4:
        start_year = year
    elif month < 4:
        start_year = year - 1
    else:
        start_year = year if salary_day >= 6 else year - 1
    return f"{start_year}/{str(start_year + 1)[-2:]}"


def _build_tax_year_projection_buckets(
    start_month,
    total_months,
    current_value,
    value_at_month_fn,
    *,
    assumptions=None,
    month_breakdown_fn=None,
    value_no_fees_at_month_fn=None,
):
    """Aggregate projection months into UK tax-year buckets.

    ISA, Lifetime ISA and pension allowances are tax-year based, so the export's
    yearly contribution rows must not make a Jan-Jul LISA fill look like one
    £7k allowance-year payment. The month-by-month sheet remains the source of
    truth; this helper groups those months into 6 Apr–5 Apr allowance years.
    """
    if not start_month or total_months <= 0:
        return []

    buckets = []
    bucket = None

    for month_count in range(1, total_months + 1):
        month_key = add_months_to_key(start_month, month_count - 1)
        tax_year = _tax_year_label_for_month_key(month_key, assumptions)
        if bucket is None or bucket["tax_year"] != tax_year:
            if bucket is not None:
                buckets.append(bucket)
            bucket = {
                "tax_year": tax_year,
                "end_month_count": month_count,
                "end_value": current_value,
                "end_value_no_fees": current_value,
                "personal": 0.0,
                "into_pot": 0.0,
                "contrib_fee": 0.0,
            }

        bucket["end_month_count"] = month_count
        bucket["end_value"] = float(value_at_month_fn(month_count) or 0)
        if value_no_fees_at_month_fn is not None:
            bucket["end_value_no_fees"] = float(value_no_fees_at_month_fn(month_count) or 0)

        if month_breakdown_fn is not None:
            breakdown = month_breakdown_fn(month_count - 1) or {}
            bucket["personal"] += float(breakdown.get("personal") or 0)
            bucket["into_pot"] += float(breakdown.get("total_into_pot") or 0)
            bucket["contrib_fee"] += float(breakdown.get("contribution_fee") or 0)

    if bucket is not None:
        buckets.append(bucket)

    return buckets


# ── Projections export ────────────────────────────────────────────────────────

@export_bp.route("/projections/export.xlsx")
@login_required
def export_projections():
    uid = current_user.id
    raw_accounts = fetch_all_accounts(uid)
    assumptions  = fetch_assumptions(uid)
    holdings_totals = fetch_holding_totals_by_account(uid)

    accounts = []
    start_month = projection_start_month_key(assumptions)
    for a in raw_accounts:
        row = dict(a)
        row["current_value"] = effective_account_value(a, holdings_totals)
        row["_contribution_overrides"] = fetch_contribution_overrides(a["id"])
        row["_projection_start_month"] = start_month
        accounts.append(row)

    current_age    = current_age_from_assumptions(assumptions) if assumptions else 43
    retirement_age = to_float(assumptions["retirement_age"]) if assumptions else 60
    growth_rate    = to_float(assumptions["annual_growth_rate"]) if assumptions else 0.07
    exact_years    = years_to_retirement(current_age, retirement_age, assumptions) if assumptions else max(retirement_age - current_age, 0)
    whole_years    = int(exact_years)
    total_projected = projected_total_retirement_value(accounts, assumptions)

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _set_col_width(ws, 1, 34)
    _set_col_width(ws, 2, 18)
    _set_col_width(ws, 3, 18)
    _set_col_width(ws, 4, 20)
    _set_col_width(ws, 5, 24)

    _title_cell(ws, 1, "SteadyPlan — Retirement Scenario Estimates", 5)
    cell = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    cell.font = _SUBTITLE_FONT

    _header_row(ws, 4, ["Account", "Current Value", "You pay monthly", "Into pots monthly", "Scenario estimate at retirement"])

    # UK pound formats
    GBP  = '£#,##0.00'
    GBP0 = '£#,##0'

    for i, acc in enumerate(accounts, 5):
        proj = projected_account_value(acc, assumptions)
        effective = projection_monthly_contribution(acc, assumptions, 0)
        first_month_override = contribution_override_for_month(acc, start_month) if start_month else None
        contribution_account = dict(acc)
        if first_month_override is not None:
            contribution_account["monthly_contribution"] = first_month_override
        personal = contribution_breakdown(contribution_account, assumptions)["personal"]
        _data_row(ws, i, [
            acc["name"],
            to_float(acc["current_value"]),
            personal,
            effective,
            proj,
        ], num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP0})

    total_row = len(accounts) + 5
    total_personal_monthly = 0.0
    for a in accounts:
        first_month_override = contribution_override_for_month(a, start_month) if start_month else None
        contribution_account = dict(a)
        if first_month_override is not None:
            contribution_account["monthly_contribution"] = first_month_override
        total_personal_monthly += contribution_breakdown(contribution_account, assumptions)["personal"]
    _data_row(ws, total_row, [
        "Total",
        sum(to_float(a["current_value"]) for a in accounts),
        total_personal_monthly,
        sum(projection_monthly_contribution(a, assumptions, 0) for a in accounts),
        total_projected,
    ], bold=True, num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP0})

    # Fee impact summary (only if any account has fees)
    total_no_fees = sum(projected_account_value_no_fees(a, assumptions) for a in accounts)
    total_fee_impact = total_no_fees - total_projected
    if total_fee_impact > 0:
        r_fee = total_row + 1
        _data_row(ws, r_fee, [
            "Lifetime cost of fees",
            "",
            "",
            "",
            total_fee_impact,
        ], bold=True, num_formats={5: GBP0})
        # Colour the fee impact value in a muted red
        ws.cell(row=r_fee, column=5).font = Font(name="Aptos", bold=True, color="DC2626", size=10)
        ws.cell(row=r_fee, column=1).font = Font(name="Aptos", bold=True, color="DC2626", size=10)

    # Assumptions block
    r = (total_row + 3) if total_fee_impact > 0 else (total_row + 2)
    ws.cell(row=r, column=1, value="Assumptions").font = _ACCENT_FONT
    for label, val in [
        ("Current age", int(current_age)),
        ("Retirement age", int(retirement_age)),
        ("Annual growth rate", f"{growth_rate*100:.1f}%"),
        ("Years to retirement", f"{exact_years:.1f}"),
    ]:
        r += 1
        ws.cell(row=r, column=1, value=label).font = _SUBTITLE_FONT
        ws.cell(row=r, column=2, value=val).font = _DATA_FONT

    r += 2
    ws.cell(row=r, column=1, value="Cash-accessible, invested-accessible, restricted, and locked-for-later money").font = _ACCENT_FONT
    _header_row(ws, r + 1, ["Type", "Current value", "Scenario estimate at retirement", "Account count"])
    access_rows = {
        "Cash accessible": [0.0, 0.0, 0],
        "Invested accessible": [0.0, 0.0, 0],
        "Restricted": [0.0, 0.0, 0],
        "Locked for later": [0.0, 0.0, 0],
    }
    for acc in accounts:
        classification = classify_account(acc)
        if classification.access_type == "accessible":
            label = classification.label
        elif classification.access_type == "restricted":
            label = "Restricted"
        else:
            label = "Locked for later"
        access_rows[label][0] += to_float(acc["current_value"])
        access_rows[label][1] += projected_account_value(acc, assumptions)
        access_rows[label][2] += 1
    for label in ["Cash accessible", "Invested accessible", "Restricted", "Locked for later"]:
        r += 1
        _data_row(ws, r + 1, [label, access_rows[label][0], access_rows[label][1], access_rows[label][2]], num_formats={2: GBP, 3: GBP0})

    r += 3
    ws.cell(row=r, column=1, value="Notes").font = _ACCENT_FONT
    for note in [
        "Values are nominal scenario estimates before inflation unless stated otherwise.",
        "You pay monthly is your personal contribution; into pots includes tax relief, employer contributions and bonuses where applicable.",
        "This is a planning estimate, not financial advice.",
    ]:
        r += 1
        ws.cell(row=r, column=1, value=note).font = _SUBTITLE_FONT

    # ── Sheet 2: Assumptions ─────────────────────────────────────────────────
    ws_ass = wb.create_sheet("Scenario Estimate Assumptions")
    _title_cell(ws_ass, 1, "SteadyPlan — Scenario Estimate Assumptions", 3)
    _set_col_width(ws_ass, 1, 32)
    _set_col_width(ws_ass, 2, 24)
    _set_col_width(ws_ass, 3, 52)
    _header_row(ws_ass, 3, ["Setting", "Value", "Note"])
    assumption_rows = [
        ("Generated", datetime.now().strftime("%d %b %Y at %H:%M"), "Snapshot date for this export."),
        ("Current age", int(current_age), "Derived from date of birth when available."),
        ("Retirement age", int(retirement_age), "Target age used for this scenario estimate."),
        ("Years to retirement", round(exact_years, 1), "Exact years when a retirement date is available."),
        ("Scenario estimate start month", start_month, "First future contribution month considered by scenario estimates."),
        ("Annual growth rate", f"{growth_rate*100:.1f}%", "Default gross annual growth before account fees."),
        ("Inflation treatment", "Nominal", "Future values are not inflation-adjusted in this export."),
        ("Salary/review day", _safe_get(assumptions, "salary_day", "") if assumptions else "", "Used to decide whether the current month has already settled."),
    ]
    for idx, row_vals in enumerate(assumption_rows, 4):
        _data_row(ws_ass, idx, row_vals)

    # ── Sheet 3: Contribution schedule ───────────────────────────────────────
    ws_sched = wb.create_sheet("Contribution Schedule")
    _title_cell(ws_sched, 1, "SteadyPlan — Contribution Schedule", 7)
    for col, width in enumerate([28, 18, 16, 16, 16, 16, 36], 1):
        _set_col_width(ws_sched, col, width)
    _header_row(ws_sched, 3, ["Account", "Wrapper", "From", "To", "You pay monthly", "Into pot monthly", "Reason"])
    sched_row = 4
    for acc in accounts:
        baseline_breakdown = contribution_breakdown(acc, assumptions)
        _data_row(ws_sched, sched_row, [
            acc["name"], acc.get("wrapper_type") or "", start_month or "Now", "Ongoing",
            baseline_breakdown["personal"], baseline_breakdown["total_into_pot"], "Account default",
        ], num_formats={5: GBP, 6: GBP})
        sched_row += 1
        for override in acc.get("_contribution_overrides", []) or []:
            adjusted = dict(acc)
            adjusted["monthly_contribution"] = override["override_amount"]
            override_month_index = None
            if start_month:
                start_idx = month_key_to_index(start_month)
                override_idx = month_key_to_index(override["from_month"])
                if start_idx is not None and override_idx is not None:
                    override_month_index = override_idx - start_idx
            if override_month_index is not None and override_month_index >= 0:
                b = projected_contribution_breakdown(acc, assumptions, override_month_index)
            else:
                b = contribution_breakdown(adjusted, assumptions)
            _data_row(ws_sched, sched_row, [
                acc["name"], acc.get("wrapper_type") or "", override["from_month"], _display_schedule_to_month(override["to_month"]),
                b["personal"], b["total_into_pot"], _safe_get(override, "reason") or "Override",
            ], num_formats={5: GBP, 6: GBP})
            sched_row += 1
        if is_premium_bonds_account(acc):
            _data_row(ws_sched, sched_row, [
                acc["name"], acc.get("wrapper_type") or "", "Cap", "", "", PREMIUM_BONDS_MAX_BALANCE,
                "NS&I Premium Bonds balance is capped; overflow is not shown as account growth.",
            ], num_formats={6: GBP0})
            sched_row += 1

    # ── Sheet 4: Year by year ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Year by Year")
    _title_cell(ws2, 1, "SteadyPlan — Tax-Year Scenario Estimate", 3)
    _header_row(ws2, 3, ["Age", "Tax year", "Scenario estimate total"])
    _set_col_width(ws2, 1, 10)
    _set_col_width(ws2, 2, 10)
    _set_col_width(ws2, 3, 22)

    curr_tax_year = uk_tax_year_label(date.today())
    total_months = int(exact_years * 12)
    current_total = sum(to_float(a["current_value"]) for a in accounts)
    _data_row(ws2, 4, [int(current_age), f"{curr_tax_year} (today)", current_total], num_formats={3: GBP0})

    portfolio_buckets = _build_tax_year_projection_buckets(
        start_month,
        total_months,
        current_total,
        lambda month_count: sum(projected_account_value_at_month(a, assumptions, month_count) for a in accounts),
        assumptions=assumptions,
    )
    for idx, bucket in enumerate(portfolio_buckets, start=5):
        age = int(current_age + bucket["end_month_count"] / 12.0)
        _data_row(ws2, idx, [age, bucket["tax_year"], bucket["end_value"]], num_formats={3: GBP0})

    # Final fractional-year point (matches summary card exactly)
    if exact_years > whole_years and (
        not portfolio_buckets or abs(float(portfolio_buckets[-1]["end_value"]) - float(total_projected)) > 0.005
    ):
        final_row = 4 + len(portfolio_buckets) + 1
        _data_row(ws2, final_row, [
            int(retirement_age),
            "Retirement",
            total_projected,
        ], bold=True, num_formats={3: GBP0})

    # ── Sheet 3: Month by month (total portfolio) ────────────────────────────
    ws3 = wb.create_sheet("Month by Month")
    _title_cell(ws3, 1, "SteadyPlan — Monthly Scenario Estimate", 3)
    _header_row(ws3, 3, ["Month", "Scenario estimate total"])
    _set_col_width(ws3, 1, 16)
    _set_col_width(ws3, 2, 22)

    for m in range(0, total_months + 1):
        month_label = "Today" if m == 0 else add_months_to_key(start_month, m - 1)
        total = sum(
            projected_account_value_at_month(a, assumptions, m)
            for a in accounts
        )
        _data_row(ws3, m + 4, [month_label, total], num_formats={2: GBP0})

    # Final row at retirement
    _data_row(ws3, total_months + 5, ["Retirement", total_projected], bold=True, num_formats={2: GBP0})

    # ── Per-account sheets: year-by-year for each account ─────────────────
    for acc in accounts:
        # Sanitise name for Excel sheet title (max 31 chars, no special chars)
        safe_name = acc["name"][:28].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "(").replace("]", ")")
        ws_acc = wb.create_sheet(safe_name)

        acc_growth = account_growth_rate(acc, assumptions)
        acc_gross = account_gross_growth_rate(acc, assumptions)
        acc_fee_pct = effective_fee_pct(acc)
        acc_platform_pct = to_float(_safe_get(acc, "platform_fee_pct", 0))
        acc_platform_flat = to_float(_safe_get(acc, "platform_fee_flat", 0))
        acc_platform_cap = to_float(_safe_get(acc, "platform_fee_cap", 0))
        acc_fund_pct = to_float(_safe_get(acc, "fund_fee_pct", 0))
        acc_contribution_fee_pct = to_float(_safe_get(acc, "contribution_fee_pct", 0))
        start_month = _safe_get(acc, "_projection_start_month")
        first_month_override = None
        if start_month:
            first_month_override = contribution_override_for_month(acc, start_month)
        contribution_account = dict(acc)
        if first_month_override is not None:
            contribution_account["monthly_contribution"] = first_month_override
        acc_monthly = projection_monthly_contribution(acc, assumptions, 0)
        acc_current = to_float(acc["current_value"])
        acc_projected = projected_account_value(acc, assumptions)
        acc_projected_no_fees = projected_account_value_no_fees(acc, assumptions)
        acc_fee_impact = acc_projected_no_fees - acc_projected
        acc_breakdown = projected_contribution_breakdown(acc, assumptions, 0)
        acc_total_months_for_fees = int(exact_years * 12)
        next_planned_month = None
        next_planned_breakdown = None
        has_annual_fees = acc_fee_pct > 0
        has_contrib_fee = acc_contribution_fee_pct > 0
        has_fees = has_annual_fees or has_contrib_fee

        is_lisa = acc.get("wrapper_type") == "Lifetime ISA"
        is_pb = is_premium_bonds_account(acc)

        def _month_breakdown(month_index):
            return projected_contribution_breakdown(acc, assumptions, month_index)

        first_contrib_fee_monthly = 0.0
        acc_total_contrib_fees = 0.0
        if acc_total_months_for_fees > 0:
            for mi in range(0, acc_total_months_for_fees):
                if is_lisa and (current_age + mi / 12.0) >= 50:
                    continue
                b = _month_breakdown(mi)
                if next_planned_month is None and (
                    abs(float(b.get("personal") or 0)) > 0.005 or abs(float(b.get("total_into_pot") or 0)) > 0.005
                ):
                    next_planned_month = add_months_to_key(start_month, mi) if start_month else None
                    next_planned_breakdown = b
        if has_contrib_fee:
            if not (is_lisa and current_age >= 50):
                first_contrib_fee_monthly = float(_month_breakdown(0).get("contribution_fee") or 0)
            for mi in range(0, acc_total_months_for_fees):
                if is_lisa and (current_age + mi / 12.0) >= 50:
                    continue
                b = _month_breakdown(mi)
                acc_total_contrib_fees += float(b.get("contribution_fee") or 0)

        max_cols = 9 if (has_annual_fees and has_contrib_fee) else (8 if has_annual_fees else (7 if has_contrib_fee or is_pb else 6))
        _title_cell(ws_acc, 1, f"SteadyPlan — {acc['name']}", max_cols)
        sub = ws_acc.cell(row=2, column=1, value=f"{acc['wrapper_type']} · {acc.get('provider') or ''}")
        sub.font = _SUBTITLE_FONT

        # Account summary
        _header_row(ws_acc, 4, ["", "Value"])
        _set_col_width(ws_acc, 1, 28)
        _set_col_width(ws_acc, 2, 18)
        summary_rows = [
            ("Current value", acc_current, GBP),
            ("You pay (monthly)", acc_breakdown["personal"], GBP),
        ]
        if (
            next_planned_month
            and next_planned_breakdown is not None
            and next_planned_month != start_month
            and (
                abs(float(next_planned_breakdown.get("personal") or 0) - float(acc_breakdown["personal"] or 0)) > 0.005
                or abs(float(next_planned_breakdown.get("total_into_pot") or 0) - float(acc_monthly or 0)) > 0.005
            )
        ):
            summary_rows.extend([
                ("Next planned month", next_planned_month, None),
                ("You pay (next planned month)", float(next_planned_breakdown.get("personal") or 0), GBP),
                ("Total into pot (next planned month)", float(next_planned_breakdown.get("total_into_pot") or 0), GBP),
            ])
        if first_month_override is not None and abs(to_float(acc.get("monthly_contribution", 0)) - to_float(first_month_override)) > 0.005:
            summary_rows.append(("Account setting (monthly)", to_float(acc.get("monthly_contribution", 0)), GBP))
        if has_contrib_fee:
            summary_rows.append(("Contribution fee deducted (monthly)", -first_contrib_fee_monthly, GBP))
        summary_rows += [
            ("Total into pot (monthly)", acc_monthly, GBP),
            ("Growth rate (net of fees)", f"{acc_growth*100:.1f}%", None),
        ]
        if has_annual_fees:
            summary_rows.append(("Growth rate (gross)", f"{acc_gross*100:.1f}%", None))
            # Show granular fee breakdown if available
            if acc_platform_pct > 0:
                cap_note = f" (capped £{acc_platform_cap:,.0f}/yr)" if acc_platform_cap > 0 else ""
                summary_rows.append(("Platform fee", f"{acc_platform_pct:.2f}%{cap_note}", None))
            if acc_platform_flat > 0:
                summary_rows.append(("Platform fee (flat)", f"£{acc_platform_flat:,.0f}/yr", None))
            if acc_fund_pct > 0:
                summary_rows.append(("Fund fee (OCF)", f"{acc_fund_pct:.2f}%", None))
            summary_rows.append(("Total effective annual fee", f"{acc_fee_pct:.2f}%", None))
        if has_contrib_fee:
            summary_rows.append(("Contribution fee", f"{acc_contribution_fee_pct:.2f}% per contribution", None))
            summary_rows.append(("Total contribution fees paid", acc_total_contrib_fees, GBP0))
        summary_rows.append(("Scenario estimate at retirement", acc_projected, GBP0))
        if has_annual_fees:
            summary_rows.append(("Value without annual fees", acc_projected_no_fees, GBP0))
            summary_rows.append(("Lifetime cost of annual fees", acc_fee_impact, GBP0))

        for ri, (label, val, fmt) in enumerate(summary_rows, 5):
            _data_row(ws_acc, ri, [label, val], num_formats={2: fmt} if fmt else {})

        # Year-by-year table — columns vary by which fees apply
        yby_start = 5 + len(summary_rows) + 1
        if has_annual_fees and has_contrib_fee:
            yby_headers = ["Age", "Tax year", "Scenario estimate value", "Growth", "You pay (yr)", "Into pot (yr)", "Contrib. Fee (yr)", "Value (no ann. fees)", "Ann. Fee Impact"]
        elif has_annual_fees:
            yby_headers = ["Age", "Tax year", "Scenario estimate value", "Growth", "You pay (yr)", "Into pot (yr)", "Value (no fees)", "Fee Impact"]
        elif has_contrib_fee:
            yby_headers = ["Age", "Tax year", "Scenario estimate value", "Growth", "You pay (yr)", "Into pot (yr)", "Contrib. Fee (yr)"]
        elif is_pb:
            yby_headers = ["Age", "Tax year", "Scenario estimate value", "Growth", "Cap adjustment", "You pay (yr)", "Into pot (yr)"]
        else:
            yby_headers = ["Age", "Tax year", "Scenario estimate value", "Growth", "You pay (yr)", "Into pot (yr)"]
        _header_row(ws_acc, yby_start, yby_headers)
        _set_col_width(ws_acc, 3, 22)
        _set_col_width(ws_acc, 4, 18)
        _set_col_width(ws_acc, 5, 18)
        if len(yby_headers) >= 6:
            _set_col_width(ws_acc, 6, 22)
        if len(yby_headers) >= 7:
            _set_col_width(ws_acc, 7, 22)
        if len(yby_headers) >= 8:
            _set_col_width(ws_acc, 8, 18)
        if len(yby_headers) >= 9:
            _set_col_width(ws_acc, 9, 18)

        prev_val = acc_current
        acc_total_months = int(exact_years * 12)
        yearly_buckets = _build_tax_year_projection_buckets(
            start_month,
            acc_total_months,
            acc_current,
            lambda month_count: projected_account_value_at_month(acc, assumptions, month_count),
            assumptions=assumptions,
            month_breakdown_fn=_month_breakdown,
            value_no_fees_at_month_fn=(
                (lambda month_count: projected_account_value_at_month_no_fees(acc, assumptions, month_count))
                if has_annual_fees else None
            ),
        )

        if has_annual_fees and has_contrib_fee:
            _data_row(ws_acc, yby_start + 1, [
                int(current_age), f"{curr_tax_year} (today)", acc_current, 0, 0, 0,
                0, acc_current, 0,
            ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0, 8: GBP0, 9: GBP0})
        elif has_annual_fees:
            _data_row(ws_acc, yby_start + 1, [
                int(current_age), f"{curr_tax_year} (today)", acc_current, 0, 0, 0,
                acc_current, 0,
            ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0, 8: GBP0})
        elif has_contrib_fee:
            _data_row(ws_acc, yby_start + 1, [
                int(current_age), f"{curr_tax_year} (today)", acc_current, 0, 0, 0, 0,
            ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0})
        elif is_pb:
            _data_row(ws_acc, yby_start + 1, [
                int(current_age), f"{curr_tax_year} (today)", acc_current, 0, 0, 0, 0,
            ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0})
        else:
            _data_row(ws_acc, yby_start + 1, [
                int(current_age), f"{curr_tax_year} (today)", acc_current, 0, 0, 0,
            ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0})

        prev_val = acc_current
        for bucket_index, bucket in enumerate(yearly_buckets, start=1):
            age = int(current_age + bucket["end_month_count"] / 12.0)
            val = bucket["end_value"]
            personal_this_year = bucket["personal"]
            contrib_this_year = bucket["into_pot"]
            contrib_fee_this_year = bucket["contrib_fee"]
            growth_this_year = val - prev_val - contrib_this_year
            cap_adjustment_this_year = 0.0
            if is_pb and growth_this_year < 0 and val >= PREMIUM_BONDS_MAX_BALANCE - 0.005:
                cap_adjustment_this_year = growth_this_year
                growth_this_year = 0.0
            year_label = bucket["tax_year"]
            row_num = yby_start + 1 + bucket_index

            if has_annual_fees and has_contrib_fee:
                val_no_fees = bucket["end_value_no_fees"]
                _data_row(ws_acc, row_num, [
                    age, year_label, val, growth_this_year, personal_this_year, contrib_this_year,
                    contrib_fee_this_year, val_no_fees, val_no_fees - val,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0, 8: GBP0, 9: GBP0})
            elif has_annual_fees:
                val_no_fees = bucket["end_value_no_fees"]
                _data_row(ws_acc, row_num, [
                    age, year_label, val, growth_this_year, personal_this_year, contrib_this_year,
                    val_no_fees, val_no_fees - val,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0, 8: GBP0})
            elif has_contrib_fee:
                _data_row(ws_acc, row_num, [
                    age, year_label, val, growth_this_year, personal_this_year, contrib_this_year, contrib_fee_this_year,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0})
            elif is_pb:
                _data_row(ws_acc, row_num, [
                    age, year_label, val, growth_this_year, cap_adjustment_this_year, personal_this_year, contrib_this_year,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0, 7: GBP0})
            else:
                _data_row(ws_acc, row_num, [
                    age, year_label, val, growth_this_year, personal_this_year, contrib_this_year,
                ], num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0})
            prev_val = val

        # Final fractional-year row
        final_year_value = float(yearly_buckets[-1]["end_value"]) if yearly_buckets else float(acc_current)
        if exact_years > whole_years and abs(final_year_value - float(acc_projected)) > 0.005:
            final_r = yby_start + 1 + len(yearly_buckets) + 1
            if has_annual_fees and has_contrib_fee:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), "Retirement", acc_projected, "", "", "", "",
                    acc_projected_no_fees, acc_fee_impact,
                ], bold=True, num_formats={3: GBP0, 8: GBP0, 9: GBP0})
            elif has_annual_fees:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), "Retirement", acc_projected, "", "", "",
                    acc_projected_no_fees, acc_fee_impact,
                ], bold=True, num_formats={3: GBP0, 7: GBP0, 8: GBP0})
            elif has_contrib_fee:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), "Retirement", acc_projected, "", "", "", "",
                ], bold=True, num_formats={3: GBP0})
            elif is_pb:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), "Retirement", acc_projected, "", "", "", "",
                ], bold=True, num_formats={3: GBP0})
            else:
                _data_row(ws_acc, final_r, [
                    int(retirement_age), "Retirement", acc_projected, "", "", "",
                ], bold=True, num_formats={3: GBP0})

        # ── Monthly breakdown table ──────────────────────────────────
        yearly_end = yby_start + 2 + len(yearly_buckets) + (1 if exact_years > whole_years and abs(final_year_value - float(acc_projected)) > 0.005 else 0)
        mby_start = yearly_end + 2  # gap of 1 empty row

        if has_annual_fees and has_contrib_fee:
            _header_row(ws_acc, mby_start, ["Month", "You pay (mo)", "Scenario estimate value", "Contrib. Fee (mo)", "Value (no ann. fees)", "Ann. Fee Impact"])
        elif has_annual_fees:
            _header_row(ws_acc, mby_start, ["Month", "You pay (mo)", "Scenario estimate value", "Value (no fees)", "Fee Impact"])
        elif has_contrib_fee:
            _header_row(ws_acc, mby_start, ["Month", "You pay (mo)", "Scenario estimate value", "Contrib. Fee (mo)"])
        else:
            _header_row(ws_acc, mby_start, ["Month", "You pay (mo)", "Scenario estimate value"])

        acc_total_months = int(exact_years * 12)
        for m in range(0, acc_total_months + 1):
            m_label = "Today" if m == 0 else add_months_to_key(start_month, m - 1)
            if m == 0:
                pass
            m_val = projected_account_value_at_month(acc, assumptions, m)
            if m == 0:
                m_personal = 0.0
                m_contrib_fee = 0.0
            else:
                mi = m - 1
                if is_lisa and (current_age + mi / 12.0) >= 50:
                    m_personal = 0.0
                    m_contrib_fee = 0.0
                else:
                    b = _month_breakdown(mi)
                    m_personal = float(b.get("personal") or 0)
                    m_contrib_fee = float(b.get("contribution_fee") or 0)

            if has_annual_fees and has_contrib_fee:
                m_val_nf = projected_account_value_at_month_no_fees(acc, assumptions, m)
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_personal, m_val, m_contrib_fee, m_val_nf, m_val_nf - m_val,
                ], num_formats={2: GBP0, 3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0})
            elif has_annual_fees:
                m_val_nf = projected_account_value_at_month_no_fees(acc, assumptions, m)
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_personal, m_val, m_val_nf, m_val_nf - m_val,
                ], num_formats={2: GBP0, 3: GBP0, 4: GBP0, 5: GBP0})
            elif has_contrib_fee:
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_personal, m_val, m_contrib_fee,
                ], num_formats={2: GBP0, 3: GBP0, 4: GBP0})
            else:
                _data_row(ws_acc, mby_start + 1 + m, [
                    m_label, m_personal, m_val,
                ], num_formats={2: GBP0, 3: GBP0})

        # Final retirement row
        m_final_r = mby_start + 1 + acc_total_months + 1
        if has_annual_fees and has_contrib_fee:
            _data_row(ws_acc, m_final_r, [
                "Retirement", "", acc_projected, acc_total_contrib_fees, acc_projected_no_fees, acc_fee_impact,
            ], bold=True, num_formats={3: GBP0, 4: GBP0, 5: GBP0, 6: GBP0})
        elif has_annual_fees:
            _data_row(ws_acc, m_final_r, [
                "Retirement", "", acc_projected, acc_projected_no_fees, acc_fee_impact,
            ], bold=True, num_formats={3: GBP0, 4: GBP0, 5: GBP0})
        elif has_contrib_fee:
            _data_row(ws_acc, m_final_r, [
                "Retirement", "", acc_projected, acc_total_contrib_fees,
            ], bold=True, num_formats={3: GBP0, 4: GBP0})
        else:
            _data_row(ws_acc, m_final_r, [
                "Retirement", "", acc_projected,
            ], bold=True, num_formats={3: GBP0})

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"projections_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Budget export ─────────────────────────────────────────────────────────────

GBP = '£#,##0.00'


def _income_key(db_sections):
    income_section = next((s for s in db_sections if "income" in s["key"].lower()), None)
    return income_section["key"] if income_section else (db_sections[0]["key"] if db_sections else "income")


def _budget_month_row_map(db_sections, items):
    row = 4
    out = {}
    for sec in db_sections:
        section_items = [it for it in items if it["section"] == sec["key"]]
        if not section_items:
            continue
        row += 1
        for it in section_items:
            out[int(it["id"])] = row
            row += 1
        row += 2
    return out


def _write_budget_month_sheet(ws, title_text, db_sections, items, entry_map, item_id_col=False, linked_accounts=None, active_overrides=None, prior_entry_map=None):
    """Render one month's budget into `ws`. Returns a dict of section_key → total.

    If item_id_col is True, an extra hidden column A carries the budget_item_id
    (used by the annual export so a future re-upload can match by ID).
    """
    col_offset = 1 if item_id_col else 0
    _set_col_width(ws, 1 + col_offset, 30)
    _set_col_width(ws, 2 + col_offset, 20)
    _set_col_width(ws, 3 + col_offset, 16)
    if item_id_col:
        _set_col_width(ws, 1, 8)
        # Actually hide column A so users don't see raw item_ids.
        # The ids are still read on re-upload to match rows reliably.
        ws.column_dimensions["A"].hidden = True

    _title_cell(ws, 1, title_text, 3 + col_offset)
    cell = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    cell.font = _SUBTITLE_FONT

    row = 4
    section_totals = {}
    section_total_cell_refs = {}
    amount_col = 3 + col_offset
    amount_col_letter = get_column_letter(amount_col)
    linked_accounts = linked_accounts or {}
    active_overrides = active_overrides or {}
    prior_entry_map = prior_entry_map or {}
    pre_salary_amount_refs = []

    for sec in db_sections:
        section_items = [it for it in items if it["section"] == sec["key"]]
        if not section_items:
            continue

        header_vals = ([""] if item_id_col else []) + [sec["label"], "", "Amount"]
        _header_row(ws, row, header_vals)
        row += 1

        sec_total = 0.0
        item_start_row = row
        for item in section_items:
            linked = linked_accounts.get(int(item["linked_account_id"])) if item.get("linked_account_id") else None
            if item["id"] in entry_map:
                amount = float(entry_map[item["id"]]["amount"] or 0)
            elif linked and int(linked["id"]) in active_overrides:
                amount = float((active_overrides[int(linked["id"])] or {}).get("override_amount") or 0)
            elif linked:
                amount = float(linked.get("monthly_contribution") or 0)
            elif item["id"] in prior_entry_map:
                amount = float(prior_entry_map[item["id"]]["amount"] or 0)
            else:
                amount = float(item["default_amount"] or 0)
            vals = ([item["id"]] if item_id_col else []) + [item["name"], item["notes"] or "", amount]
            _data_row(ws, row, vals, num_formats={3 + col_offset: GBP})
            if linked and int(linked.get("pre_salary") or 0):
                pre_salary_amount_refs.append(f"{amount_col_letter}{row}")
            sec_total += amount
            row += 1
        item_end_row = row - 1

        total_formula = f"=SUM({amount_col_letter}{item_start_row}:{amount_col_letter}{item_end_row})" if item_end_row >= item_start_row else "=0"
        total_vals = ([""] if item_id_col else []) + ["", "Section total", total_formula]
        _data_row(ws, row, total_vals, bold=True, num_formats={3 + col_offset: GBP})
        section_totals[sec["key"]] = sec_total
        section_total_cell_refs[sec["key"]] = f"{amount_col_letter}{row}"
        row += 2

    total_income = section_totals.get(_income_key(db_sections), 0)
    total_expenses = sum(v for k, v in section_totals.items() if k != _income_key(db_sections))
    surplus = total_income - total_expenses

    income_key = _income_key(db_sections)
    income_ref = section_total_cell_refs.get(income_key)
    expense_refs = [ref for k, ref in section_total_cell_refs.items() if k != income_key]

    total_income_formula = f"={income_ref}" if income_ref else "=0"
    total_expenses_formula = f"=SUM({', '.join(expense_refs)})" if expense_refs else "=0"

    income_row = row
    income_cell_ref = f"{amount_col_letter}{income_row}"
    row_vals = ([""] if item_id_col else []) + ["Total Income", "", total_income_formula]
    _data_row(ws, row, row_vals, num_formats={3 + col_offset: GBP})
    row += 1

    expenses_row = row
    expenses_cell_ref = f"{amount_col_letter}{expenses_row}"
    row_vals = ([""] if item_id_col else []) + ["Total Expenses", "", total_expenses_formula]
    _data_row(ws, row, row_vals, num_formats={3 + col_offset: GBP})
    row += 1

    pre_salary_row = row
    pre_salary_cell_ref = f"{amount_col_letter}{pre_salary_row}"
    pre_salary_formula = f"=SUM({', '.join(pre_salary_amount_refs)})" if pre_salary_amount_refs else "=0"
    row_vals = ([""] if item_id_col else []) + ["Outside take-home (not from pocket)", "", pre_salary_formula]
    _data_row(ws, row, row_vals, num_formats={3 + col_offset: GBP})
    row += 1

    row_vals = ([""] if item_id_col else []) + ["Surplus", "", f"={income_cell_ref}-{expenses_cell_ref}+{pre_salary_cell_ref}"]
    _data_row(ws, row, row_vals, bold=True, num_formats={3 + col_offset: GBP})
    row += 1

    return section_totals


@export_bp.route("/budget/export.xlsx")
@login_required
def export_budget():
    uid = current_user.id
    month_key = valid_month_key(request.args.get("month")) or date.today().strftime("%Y-%m")
    month_label = datetime.strptime(month_key, "%Y-%m").strftime("%B %Y")

    db_sections = fetch_budget_sections(uid)
    items = fetch_budget_items(uid)
    accounts = fetch_all_accounts(uid)
    account_map = {int(a["id"]): dict(a) for a in accounts}
    entries = fetch_budget_entries(month_key, uid)
    entry_map = {e["budget_item_id"]: e for e in entries}
    prior = fetch_prior_month_budget_entries(month_key, uid)
    prior_entry_map = {e["budget_item_id"]: e for e in prior}
    active_overrides = fetch_all_active_overrides(month_key, uid) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {month_key}"
    _write_budget_month_sheet(
        ws,
        f"SteadyPlan — Budget for {month_label}",
        db_sections,
        items,
        entry_map,
        linked_accounts=account_map,
        active_overrides=active_overrides,
        prior_entry_map=prior_entry_map,
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"budget_{month_key}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Annual budget export (UK tax year) ────────────────────────────────────────

_MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _tax_year_months(start_year):
    """Return the 12 month_key strings ('YYYY-MM') covering tax year start_year/start_year+1.

    UK tax year: Apr (start_year) through Mar (start_year + 1).
    """
    result = []
    for offset in range(12):
        m = 4 + offset
        y = start_year + (m - 1) // 12
        month = ((m - 1) % 12) + 1
        result.append(f"{y:04d}-{month:02d}")
    return result


def _resolved_month_map(month_key, uid, carry_forward):
    """Entry map for a month. If the month has no entries, fall back to the most
    recent prior month in-workbook (carry_forward), then to DB prior-month, then
    to default_amount (handled by the sheet writer)."""
    entries = fetch_budget_entries(month_key, uid)
    entry_map = {e["budget_item_id"]: e for e in entries}
    if not entry_map and carry_forward:
        # Reuse the last month we already resolved so the year stays consistent
        return dict(carry_forward)
    if not entry_map:
        prior = fetch_prior_month_budget_entries(month_key, uid)
        entry_map = {e["budget_item_id"]: e for e in prior}
    return entry_map


def _write_annual_summary_sheet(ws, months, month_labels, db_sections, items, month_entry_maps, pre_salary_item_ids=None):
    n_months = len(months)
    total_col = 4 + n_months
    _set_col_width(ws, 1, 8)
    _set_col_width(ws, 2, 30)
    _set_col_width(ws, 3, 20)
    for col in range(4, total_col):
        _set_col_width(ws, col, 12)
    _set_col_width(ws, total_col, 14)
    ws.column_dimensions["A"].hidden = True

    _title_cell(ws, 1, "SteadyPlan — Annual Budget Summary", total_col)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}").font = _SUBTITLE_FONT
    note = ws.cell(
        row=3,
        column=1,
        value="How to edit: change values in the monthly tabs (e.g. “May 2026”) in the Amount column. This Summary is calculated — clicking a cell will show a formula like ='May 2026'!$D$14.",
    )
    note.font = _SUBTITLE_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_col)

    _header_row(ws, 4, ["Item ID", "Item", "Notes"] + month_labels + ["Total"])
    month_row_map = _budget_month_row_map(db_sections, items)
    pre_salary_item_ids = set(int(x) for x in (pre_salary_item_ids or []))
    pre_salary_item_rows = []

    month_start_col = 4
    month_end_col = 3 + n_months
    month_start_letter = get_column_letter(month_start_col)
    month_end_letter = get_column_letter(month_end_col)

    row = 5
    section_total_rows = {}
    num_formats = {c: GBP for c in range(month_start_col, total_col + 1)}

    for sec in db_sections:
        section_items = [it for it in items if it["section"] == sec["key"]]
        if not section_items:
            continue

        _header_row(ws, row, ["", sec["label"], ""] + [""] * n_months + [""])
        row += 1

        item_start_row = row
        for item in section_items:
            item_id = int(item["id"])
            src_row = month_row_map.get(item_id)
            month_cells = []
            for sheet_name in month_labels:
                month_cells.append(f"='{sheet_name}'!$D${src_row}" if src_row else "0")
            total_formula = f"=SUM({month_start_letter}{row}:{month_end_letter}{row})"
            _data_row(
                ws,
                row,
                [item_id, item["name"], item["notes"] or ""] + month_cells + [total_formula],
                num_formats=num_formats,
            )
            if item_id in pre_salary_item_ids:
                pre_salary_item_rows.append(row)
            row += 1
        item_end_row = row - 1

        month_totals = []
        for c in range(month_start_col, month_end_col + 1):
            col_letter = get_column_letter(c)
            month_totals.append(
                f"=SUM({col_letter}{item_start_row}:{col_letter}{item_end_row})" if item_end_row >= item_start_row else "=0"
            )
        total_formula = f"=SUM({get_column_letter(total_col)}{item_start_row}:{get_column_letter(total_col)}{item_end_row})" if item_end_row >= item_start_row else "=0"
        _data_row(ws, row, ["", "Section total", ""] + month_totals + [total_formula], bold=True, num_formats=num_formats)
        section_total_rows[sec["key"]] = row
        row += 2

    income_key = _income_key(db_sections)
    income_total_row = section_total_rows.get(income_key)
    expense_total_rows = [r for k, r in section_total_rows.items() if k != income_key]

    income_row = row
    income_cells = []
    for c in range(month_start_col, month_end_col + 1):
        col_letter = get_column_letter(c)
        income_cells.append(f"={col_letter}{income_total_row}" if income_total_row else "=0")
    income_total = f"=SUM({month_start_letter}{income_row}:{month_end_letter}{income_row})"
    _data_row(ws, income_row, ["", "Total Income", ""] + income_cells + [income_total], num_formats=num_formats)
    row += 1

    expense_row = row
    expense_cells = []
    for c in range(month_start_col, month_end_col + 1):
        col_letter = get_column_letter(c)
        refs = [f"{col_letter}{r}" for r in expense_total_rows]
        expense_cells.append(f"=SUM({', '.join(refs)})" if refs else "=0")
    expense_total = f"=SUM({month_start_letter}{expense_row}:{month_end_letter}{expense_row})"
    _data_row(ws, expense_row, ["", "Total Expenses", ""] + expense_cells + [expense_total], num_formats=num_formats)
    row += 1

    outside_row = row
    outside_cells = []
    for c in range(month_start_col, month_end_col + 1):
        col_letter = get_column_letter(c)
        refs = [f"{col_letter}{r}" for r in pre_salary_item_rows]
        outside_cells.append(f"=SUM({', '.join(refs)})" if refs else "=0")
    outside_total = f"=SUM({month_start_letter}{outside_row}:{month_end_letter}{outside_row})"
    _data_row(ws, outside_row, ["", "Outside take-home (not from pocket)", ""] + outside_cells + [outside_total], num_formats=num_formats)
    row += 1

    surplus_row = row
    surplus_cells = []
    for c in range(month_start_col, month_end_col + 1):
        col_letter = get_column_letter(c)
        surplus_cells.append(f"={col_letter}{income_row}-{col_letter}{expense_row}+{col_letter}{outside_row}")
    surplus_total = f"=SUM({month_start_letter}{surplus_row}:{month_end_letter}{surplus_row})"
    _data_row(ws, surplus_row, ["", "Surplus", ""] + surplus_cells + [surplus_total], bold=True, num_formats=num_formats)


def _write_budget_export_guide_sheet(ws, start_year):
    _set_col_width(ws, 1, 110)
    ty_label = f"{start_year}/{str(start_year + 1)[-2:]}"
    _title_cell(ws, 1, f"SteadyPlan — Export Guide (Tax Year {ty_label})", 1)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}").font = _SUBTITLE_FONT

    lines = [
        "Workflow (recommended):",
        "1) Edit numbers in the monthly tabs (Apr … Mar). Use the Amount column — those are the inputs.",
        "2) The Summary tab is calculated from the month tabs. It updates automatically when you edit a month.",
        "3) Investment Tracking shows (a) what your budget plans to contribute and (b) what you actually logged in SteadyPlan (ISA/Pension top-ups).",
        "4) Rows marked as Outside take-home pay (e.g. cashback, salary sacrifice) are added back in Surplus so they don’t reduce take-home cashflow.",
        "",
        "If you click a cell in Summary and see something like ='May 2026'!$D$14:",
        "- That’s normal: it means the value is linked to the May sheet.",
        "- If you overwrite that formula, it becomes a manual value and will no longer stay in sync with the monthly sheets.",
        "",
        "Legend (Investment Tracking):",
        "- Personal: money you plan to pay in (from budget lines linked to accounts).",
        "- Tax relief: added for relief-at-source pensions/SIPPs (25% uplift on net; logged pension rows are treated as gross).",
        "- Lifetime ISA bonus: 25% bonus on Lifetime ISA personal contributions until the £4,000/year cap (bonus does not count toward ISA allowance).",
        "- Employer: workplace pension employer contributions.",
        "- Allowance basis: ISA uses personal only; Pension uses gross into-pot (personal + relief + employer).",
    ]
    r = 4
    for t in lines:
        cell = ws.cell(row=r, column=1, value=t)
        cell.font = _DATA_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def _write_investment_tracking_sheet(ws, uid, start_year, accounts, items, month_entry_maps, assumptions):
    _set_col_width(ws, 1, 26)
    _set_col_width(ws, 2, 18)
    _set_col_width(ws, 3, 16)
    _set_col_width(ws, 4, 16)
    _set_col_width(ws, 5, 16)
    _set_col_width(ws, 6, 16)
    _set_col_width(ws, 7, 16)
    _set_col_width(ws, 8, 16)
    _set_col_width(ws, 9, 16)
    _set_col_width(ws, 10, 16)
    _set_col_width(ws, 11, 16)
    _set_col_width(ws, 12, 16)
    _set_col_width(ws, 13, 18)
    _set_col_width(ws, 14, 18)
    _set_col_width(ws, 15, 16)

    ty_label = f"{start_year}/{str(start_year + 1)[-2:]}"
    _title_cell(ws, 1, f"SteadyPlan — Investment Tracking (Tax Year {ty_label})", 15)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}").font = _SUBTITLE_FONT

    isa_allowance = float(assumptions["isa_allowance"]) if assumptions and assumptions.get("isa_allowance") else 20000.0
    lisa_allowance = float(assumptions["lisa_allowance"]) if assumptions and assumptions.get("lisa_allowance") else 4000.0
    pension_allowance = float(assumptions["pension_annual_allowance"]) if assumptions and assumptions.get("pension_annual_allowance") else 60000.0

    month_keys = list(month_entry_maps.keys())
    month_sheet_names = []
    for mk in month_keys:
        d = datetime.strptime(mk, "%Y-%m")
        month_sheet_names.append(f"{_MONTH_NAMES[d.month - 1]} {d.year}")

    db_sections = fetch_budget_sections(uid)
    item_row_map = _budget_month_row_map(db_sections, items)

    linked_items = [it for it in items if it.get("linked_account_id")]
    account_to_item_ids = {}
    for it in linked_items:
        aid = int(it["linked_account_id"])
        account_to_item_ids.setdefault(aid, []).append(int(it["id"]))

    ty_start_iso = date(start_year, 4, 6).isoformat()
    ty_end_iso = date(start_year + 1, 4, 5).isoformat()
    isa_logs = fetch_isa_contributions(uid, ty_start_iso, ty_end_iso) or []
    pension_logs = fetch_pension_contributions(uid, ty_start_iso, ty_end_iso) or []

    isa_logged_by_account = {}
    for r in isa_logs:
        aid = int(r["account_id"])
        isa_logged_by_account[aid] = isa_logged_by_account.get(aid, 0.0) + float(r["amount"] or 0)

    pension_logged_personal_gross_by_account = {}
    pension_logged_employer_by_account = {}
    for r in pension_logs:
        aid = int(r["account_id"])
        kind = (r.get("kind") or "personal").strip().lower()
        amt = float(r["amount"] or 0)
        if kind == "employer":
            pension_logged_employer_by_account[aid] = pension_logged_employer_by_account.get(aid, 0.0) + amt
        else:
            pension_logged_personal_gross_by_account[aid] = pension_logged_personal_gross_by_account.get(aid, 0.0) + amt

    def _sum_formula(refs):
        refs = [r for r in refs if r]
        if not refs:
            return "0"
        if len(refs) == 1:
            return f"={refs[0]}"
        return f"=SUM({', '.join(refs)})"

    def _planned_personal_formula_for_account(aid):
        item_ids = account_to_item_ids.get(aid) or []
        refs = []
        for item_id in item_ids:
            src_row = item_row_map.get(int(item_id))
            if not src_row:
                continue
            for sheet_name in month_sheet_names:
                refs.append(f"'{sheet_name}'!$D${src_row}")
        return _sum_formula(refs)

    account_map = {int(a["id"]): dict(a) for a in accounts}

    wrapper_table_start = 4
    _header_row(ws, wrapper_table_start, ["Wrapper", "Planned (allowance basis)", "Logged (allowance basis)", "Allowance", "Remaining", "% Used"])
    PCT = '0.0"%"'

    wrapper_rows = [
        ("ISA (all)", isa_allowance),
        ("  of which Lifetime ISA", lisa_allowance),
        ("Pension", pension_allowance),
    ]
    for i, (label, allowance) in enumerate(wrapper_rows):
        r = wrapper_table_start + 1 + i
        remaining_formula = f"=MAX($D{r}-MAX($B{r},$C{r}),0)"
        pct_formula = f"=IF($D{r}>0,MAX($B{r},$C{r})/$D{r},0)"
        _data_row(ws, r, [label, 0, 0, float(allowance), remaining_formula, pct_formula], num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP, 6: PCT})

    note_row = wrapper_table_start + 5
    note = ws.cell(row=note_row, column=1, value="Note: Lifetime ISA personal contributions count toward the overall ISA £20k allowance. The 25% Lifetime ISA bonus does not.")
    note.font = _SUBTITLE_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=15)

    row = note_row + 2
    per_account_header_row = row
    _header_row(ws, per_account_header_row, [
        "Account",
        "Wrapper",
        "Planned personal (budget)",
        "Planned tax relief",
        "Planned Lifetime ISA bonus",
        "Planned employer",
        "Planned into pot",
        "Logged personal",
        "Logged tax relief",
        "Logged Lifetime ISA bonus",
        "Logged employer",
        "Logged into pot",
        "Allowance basis (planned)",
        "Allowance basis (logged)",
        "Diff (Logged − Planned)",
    ])
    row += 1

    isa_acc_rows = []
    lisa_acc_rows = []
    pension_acc_rows = []
    for aid, acc in sorted(account_map.items(), key=lambda kv: kv[1].get("name") or ""):
        has_budget = bool(account_to_item_ids.get(aid))
        has_isa_logs = float(isa_logged_by_account.get(aid, 0.0) or 0) > 0
        has_pension_logs = float(pension_logged_personal_gross_by_account.get(aid, 0.0) or 0) > 0 or float(pension_logged_employer_by_account.get(aid, 0.0) or 0) > 0
        has_employer_plan = float(acc.get("employer_contribution") or 0) > 0
        if not (has_budget or has_isa_logs or has_pension_logs or has_employer_plan):
            continue

        wrapper = (acc.get("wrapper_type") or "").strip()
        is_isa = wrapper in ISA_WRAPPER_TYPES
        is_lisa = wrapper in LISA_WRAPPER_TYPES
        is_pension = is_pension_account(acc)
        method = (acc.get("contribution_method") or "standard").strip().lower()
        relief_at_source = ("SIPP" in wrapper) or (is_pension and method != "salary_sacrifice")

        planned_personal_cell = f"C{row}"
        planned_personal = _planned_personal_formula_for_account(aid)
        planned_tax_relief = f"={planned_personal_cell}*0.25" if relief_at_source else 0
        planned_lisa_bonus = f"=MIN({planned_personal_cell},$D{wrapper_table_start + 2})*0.25" if is_lisa else 0
        planned_employer = (float(acc.get("employer_contribution") or 0) * 12.0) if is_pension else 0.0
        planned_into_pot = f"={planned_personal_cell}+D{row}+E{row}+F{row}"

        logged_isa_personal = float(isa_logged_by_account.get(aid, 0.0) or 0.0)
        pension_gross = float(pension_logged_personal_gross_by_account.get(aid, 0.0) or 0.0)
        pension_employer = float(pension_logged_employer_by_account.get(aid, 0.0) or 0.0)

        logged_personal = logged_isa_personal
        logged_tax_relief = 0.0
        logged_employer = 0.0
        if is_pension:
            if relief_at_source and pension_gross > 0:
                logged_personal = pension_gross * 0.8
                logged_tax_relief = pension_gross * 0.2
            else:
                logged_personal = pension_gross
                logged_tax_relief = 0.0
            logged_employer = pension_employer

        logged_personal_cell = f"H{row}"
        logged_lisa_bonus = f"=MIN({logged_personal_cell},$D{wrapper_table_start + 2})*0.25" if is_lisa else 0
        logged_into_pot = f"={logged_personal_cell}+I{row}+J{row}+K{row}"

        allowance_basis_planned = f"=G{row}" if is_pension else (f"={planned_personal_cell}" if is_isa else 0)
        allowance_basis_logged = f"=L{row}" if is_pension else (f"={logged_personal_cell}" if is_isa else 0)
        diff_formula = f"=N{row}-M{row}"

        _data_row(
            ws,
            row,
            [
                acc.get("name") or "",
                wrapper or "—",
                planned_personal,
                planned_tax_relief,
                planned_lisa_bonus,
                planned_employer,
                planned_into_pot,
                logged_personal,
                logged_tax_relief,
                logged_lisa_bonus,
                logged_employer,
                logged_into_pot,
                allowance_basis_planned,
                allowance_basis_logged,
                diff_formula,
            ],
            num_formats={
                3: GBP, 4: GBP, 5: GBP, 6: GBP, 7: GBP,
                8: GBP, 9: GBP, 10: GBP, 11: GBP, 12: GBP,
                13: GBP, 14: GBP, 15: GBP,
            },
        )
        if is_isa:
            isa_acc_rows.append(row)
        if is_lisa:
            lisa_acc_rows.append(row)
        if is_pension:
            pension_acc_rows.append(row)
        row += 1

    def _sum_col(col_letter, rows_):
        refs = [f"{col_letter}{r}" for r in rows_]
        return _sum_formula(refs)

    ws.cell(row=wrapper_table_start + 1, column=2, value=_sum_col("M", isa_acc_rows)).number_format = GBP
    ws.cell(row=wrapper_table_start + 1, column=3, value=_sum_col("N", isa_acc_rows)).number_format = GBP
    ws.cell(row=wrapper_table_start + 2, column=2, value=_sum_col("M", lisa_acc_rows)).number_format = GBP
    ws.cell(row=wrapper_table_start + 2, column=3, value=_sum_col("N", lisa_acc_rows)).number_format = GBP
    ws.cell(row=wrapper_table_start + 3, column=2, value=_sum_col("M", pension_acc_rows)).number_format = GBP
    ws.cell(row=wrapper_table_start + 3, column=3, value=_sum_col("N", pension_acc_rows)).number_format = GBP

    row += 2
    _header_row(ws, row, ["Monthly totals — what each column includes", "", "", "", "", "", ""])
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    info = ws.cell(
        row=row,
        column=1,
        value="Personal is the sum of your linked budget lines (by account). Tax relief is 25% of the pension/SIPP lines that are relief-at-source. Lifetime ISA bonus is calculated with a running £4k/year cap. Employer is your pension employer contribution (or logged employer payments).",
    )
    info.font = _SUBTITLE_FONT
    info.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 2

    _header_row(ws, row, ["Account", "Wrapper", "Budget source (example)", "Personal", "Tax relief", "Lifetime ISA bonus", "Employer/mo"])
    row += 1
    item_name_map = {int(it["id"]): (it.get("name") or "") for it in items}
    first_month_sheet = month_sheet_names[0] if month_sheet_names else ""
    for aid, acc in sorted(account_map.items(), key=lambda kv: kv[1].get("name") or ""):
        item_ids = account_to_item_ids.get(aid) or []
        wrapper = (acc.get("wrapper_type") or "").strip()
        is_pension = is_pension_account(acc)
        method = (acc.get("contribution_method") or "standard").strip().lower()
        relief_at_source = ("SIPP" in wrapper) or (is_pension and method != "salary_sacrifice")
        is_lisa = wrapper in LISA_WRAPPER_TYPES
        employer_m = float(acc.get("employer_contribution") or 0.0) if is_pension else 0.0

        if not item_ids and employer_m <= 0 and float(isa_logged_by_account.get(aid, 0.0) or 0) <= 0 and float(pension_logged_personal_gross_by_account.get(aid, 0.0) or 0) <= 0 and float(pension_logged_employer_by_account.get(aid, 0.0) or 0) <= 0:
            continue

        example = "—"
        if item_ids and first_month_sheet:
            parts = []
            for item_id in item_ids[:2]:
                src_row = item_row_map.get(int(item_id))
                nm = item_name_map.get(int(item_id), "") or f"Item {item_id}"
                if src_row:
                    parts.append(f"{nm}: '{first_month_sheet}'!D{src_row}")
                else:
                    parts.append(nm)
            if len(item_ids) > 2:
                parts.append(f"+{len(item_ids) - 2} more")
            example = "; ".join(parts)

        _data_row(
            ws,
            row,
            [
                acc.get("name") or "",
                wrapper or "—",
                example,
                "Yes" if item_ids else "—",
                "25% of personal" if relief_at_source else "—",
                "25% (cap £4k/yr)" if is_lisa else "—",
                employer_m if employer_m > 0 else "—",
            ],
            num_formats={7: GBP},
        )
        row += 1

    row += 1
    _header_row(ws, row, ["Month", "Personal", "Tax relief", "Lifetime ISA bonus", "Employer", "Total into pot", "Running total"])
    row += 1

    ws.column_dimensions["P"].hidden = True
    ws.column_dimensions["Q"].hidden = True

    pension_employer_monthly = 0.0
    for acc in account_map.values():
        if is_pension_account(acc):
            pension_employer_monthly += float(acc.get("employer_contribution") or 0.0)

    lisa_item_ids = []
    pension_relief_item_ids = []
    all_linked_item_ids = [int(it["id"]) for it in linked_items]
    for it in linked_items:
        aid = int(it["linked_account_id"])
        acc = account_map.get(aid, {})
        wrapper = (acc.get("wrapper_type") or "").strip()
        is_pension = is_pension_account(acc)
        method = (acc.get("contribution_method") or "standard").strip().lower()
        relief_at_source = ("SIPP" in wrapper) or (is_pension and method != "salary_sacrifice")
        if wrapper in LISA_WRAPPER_TYPES:
            lisa_item_ids.append(int(it["id"]))
        if relief_at_source:
            pension_relief_item_ids.append(int(it["id"]))

    planned_running_total_cell = None
    lisa_running_personal_cell = None
    for idx, (mk, sheet_name) in enumerate(zip(month_keys, month_sheet_names)):
        r = row + idx
        month_label = datetime.strptime(mk, "%Y-%m").strftime("%b %Y")

        personal_refs = []
        for item_id in all_linked_item_ids:
            src_row = item_row_map.get(int(item_id))
            if src_row:
                personal_refs.append(f"'{sheet_name}'!$D${src_row}")
        personal_formula = _sum_formula(personal_refs)

        tax_refs = []
        for item_id in pension_relief_item_ids:
            src_row = item_row_map.get(int(item_id))
            if src_row:
                tax_refs.append(f"'{sheet_name}'!$D${src_row}")
        tax_formula = f"=0.25*SUM({', '.join(tax_refs)})" if tax_refs else "0"

        lisa_refs = []
        for item_id in lisa_item_ids:
            src_row = item_row_map.get(int(item_id))
            if src_row:
                lisa_refs.append(f"'{sheet_name}'!$D${src_row}")
        lisa_personal_formula = _sum_formula(lisa_refs)

        lisa_personal_cell = f"P{r}"
        lisa_running_cell = f"Q{r}"
        if idx == 0:
            lisa_running_formula = f"={lisa_personal_cell}"
        else:
            lisa_running_formula = f"={lisa_running_personal_cell}+{lisa_personal_cell}"

        bonus_formula = f"=MIN(MAX($D{wrapper_table_start + 2}-IF({lisa_running_cell}-{lisa_personal_cell}<0,0,{lisa_running_cell}-{lisa_personal_cell}),0),{lisa_personal_cell})*0.25"

        total_formula = f"=B{r}+C{r}+D{r}+E{r}"
        if planned_running_total_cell:
            running_formula = f"={planned_running_total_cell}+F{r}"
        else:
            running_formula = f"=F{r}"

        _data_row(
            ws,
            r,
            [month_label, personal_formula, tax_formula, bonus_formula, pension_employer_monthly, total_formula, running_formula],
            num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP, 6: GBP, 7: GBP},
        )
        ws.cell(row=r, column=16, value=lisa_personal_formula).number_format = GBP
        ws.cell(row=r, column=17, value=lisa_running_formula).number_format = GBP

        planned_running_total_cell = f"G{r}"
        lisa_running_personal_cell = lisa_running_cell

    row = row + len(month_keys) + 2
    _header_row(ws, row, ["Month (logged)", "Personal", "Tax relief", "Lifetime ISA bonus", "Employer", "Total into pot", "Running total"])
    row += 1

    isa_personal_by_month = {mk: 0.0 for mk in month_keys}
    lisa_personal_by_month_logged = {mk: 0.0 for mk in month_keys}
    pension_personal_net_by_month = {mk: 0.0 for mk in month_keys}
    pension_tax_relief_by_month = {mk: 0.0 for mk in month_keys}
    pension_employer_by_month = {mk: 0.0 for mk in month_keys}

    for rlog in isa_logs:
        ds = str(rlog.get("contribution_date") or "")[:10]
        try:
            d = datetime.strptime(ds, "%Y-%m-%d").date()
        except Exception:
            continue
        mk = f"{d.year:04d}-{d.month:02d}"
        if mk not in isa_personal_by_month:
            continue
        aid = int(rlog["account_id"])
        amt = float(rlog["amount"] or 0)
        isa_personal_by_month[mk] += amt
        acc = account_map.get(aid, {})
        if (acc.get("wrapper_type") or "") in LISA_WRAPPER_TYPES:
            lisa_personal_by_month_logged[mk] += amt

    for rlog in pension_logs:
        ds = str(rlog.get("contribution_date") or "")[:10]
        try:
            d = datetime.strptime(ds, "%Y-%m-%d").date()
        except Exception:
            continue
        mk = f"{d.year:04d}-{d.month:02d}"
        if mk not in pension_employer_by_month:
            continue
        aid = int(rlog["account_id"])
        acc = account_map.get(aid, {})
        wrapper = (acc.get("wrapper_type") or "").strip()
        is_pension = is_pension_account(acc)
        method = (acc.get("contribution_method") or "standard").strip().lower()
        relief_at_source = ("SIPP" in wrapper) or (is_pension and method != "salary_sacrifice")
        kind = (rlog.get("kind") or "personal").strip().lower()
        amt = float(rlog["amount"] or 0)
        if kind == "employer":
            pension_employer_by_month[mk] += amt
        else:
            if relief_at_source:
                pension_personal_net_by_month[mk] += amt * 0.8
                pension_tax_relief_by_month[mk] += amt * 0.2
            else:
                pension_personal_net_by_month[mk] += amt

    logged_running_total_cell = None
    logged_lisa_running_cell = None
    for idx, mk in enumerate(month_keys):
        r = row + idx
        month_label = datetime.strptime(mk, "%Y-%m").strftime("%b %Y")
        personal_val = float(isa_personal_by_month.get(mk, 0.0) or 0.0) + float(pension_personal_net_by_month.get(mk, 0.0) or 0.0)
        tax_val = float(pension_tax_relief_by_month.get(mk, 0.0) or 0.0)
        employer_val = float(pension_employer_by_month.get(mk, 0.0) or 0.0)

        lisa_personal_cell = f"P{r}"
        lisa_running_cell = f"Q{r}"
        ws.cell(row=r, column=16, value=float(lisa_personal_by_month_logged.get(mk, 0.0) or 0.0)).number_format = GBP
        if idx == 0:
            ws.cell(row=r, column=17, value=f"={lisa_personal_cell}").number_format = GBP
        else:
            ws.cell(row=r, column=17, value=f"={logged_lisa_running_cell}+{lisa_personal_cell}").number_format = GBP

        bonus_formula = f"=MIN(MAX($D{wrapper_table_start + 2}-IF({lisa_running_cell}-{lisa_personal_cell}<0,0,{lisa_running_cell}-{lisa_personal_cell}),0),{lisa_personal_cell})*0.25"
        total_formula = f"=B{r}+C{r}+D{r}+E{r}"
        if logged_running_total_cell:
            running_formula = f"={logged_running_total_cell}+F{r}"
        else:
            running_formula = f"=F{r}"

        _data_row(
            ws,
            r,
            [month_label, personal_val, tax_val, bonus_formula, employer_val, total_formula, running_formula],
            num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP, 6: GBP, 7: GBP},
        )

        logged_running_total_cell = f"G{r}"
        logged_lisa_running_cell = lisa_running_cell


@export_bp.route("/budget/annual-export.xlsx")
@login_required
def export_budget_annual():
    """Annual budget export: 12 month tabs (Apr→Mar of UK tax year) + Summary
    + Investment Tracking."""
    uid = current_user.id
    today = date.today()
    default_start = uk_tax_year_start(today).year
    try:
        start_year = int(request.args.get("tax_year_start") or default_start)
    except (ValueError, TypeError):
        start_year = default_start

    db_sections = fetch_budget_sections(uid)
    items = fetch_budget_items(uid)
    accounts = fetch_all_accounts(uid)
    assumptions = fetch_assumptions(uid)
    account_map = {int(a["id"]): dict(a) for a in accounts}

    months = _tax_year_months(start_year)
    month_labels = []
    for mk in months:
        d = datetime.strptime(mk, "%Y-%m")
        month_labels.append(f"{_MONTH_NAMES[d.month - 1]} {d.year}")

    # Resolve per-month entry maps (with in-workbook carry-forward for empty months)
    month_entry_maps = {}
    carry = None
    for mk in months:
        em = _resolved_month_map(mk, uid, carry)
        month_entry_maps[mk] = em
        if em:
            carry = em

    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "Guide"
    _write_budget_export_guide_sheet(ws_guide, start_year)

    ws_sum = wb.create_sheet("Summary")
    pre_salary_item_ids = {
        int(it["id"])
        for it in items
        if it.get("linked_account_id")
        and int(it["linked_account_id"]) in account_map
        and int(account_map[int(it["linked_account_id"])].get("pre_salary") or 0)
    }
    _write_annual_summary_sheet(ws_sum, months, month_labels, db_sections, items, month_entry_maps, pre_salary_item_ids=pre_salary_item_ids)

    # 12 month sheets (re-uses the monthly format with hidden item_id column A)
    for mk, label in zip(months, month_labels):
        ws = wb.create_sheet(label)
        active_overrides = fetch_all_active_overrides(mk, uid) or {}
        _write_budget_month_sheet(ws, f"SteadyPlan — Budget for {label}", db_sections, items,
                                  month_entry_maps[mk], item_id_col=True,
                                  linked_accounts=account_map,
                                  active_overrides=active_overrides)

    # Investment Tracking
    ws_inv = wb.create_sheet("Investment Tracking")
    _write_investment_tracking_sheet(ws_inv, uid, start_year, accounts, items, month_entry_maps, assumptions)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"budget_tax_year_{start_year}-{str(start_year + 1)[-2:]}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Performance export ────────────────────────────────────────────────────────

@export_bp.route("/performance/export.xlsx")
@login_required
def export_performance():
    uid = current_user.id
    current_month_key = date.today().strftime("%Y-%m")
    refresh_account_snapshots_for_month(uid, current_month_key, require_existing_month=True)
    assumptions = fetch_assumptions(uid)
    accounts = fetch_all_accounts(uid)
    account_id = request.args.get("account_id")
    selected_period = _normalise_performance_export_period(request.args.get("period"))

    assumed_rate = to_float(assumptions["annual_growth_rate"]) if assumptions else 0.07
    assumed_monthly_total = sum(to_float(a["monthly_contribution"]) for a in accounts)

    per_account_data = fetch_monthly_performance_data_by_account(uid)
    account_map = {int(a["id"]): a for a in accounts}

    selected_account_id = None
    if account_id:
        try:
            selected_account_id = int(account_id)
        except Exception:
            selected_account_id = None
        if selected_account_id not in account_map:
            selected_account_id = None

    perf_portfolio = None
    if selected_account_id is None:
        monthly_data = fetch_monthly_performance_data(uid)
        _, monthly_data = _filter_monthly_data_for_period(monthly_data, selected_period)
        perf_portfolio = compute_performance_series(monthly_data, assumed_rate, assumed_monthly_total)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _set_col_width(ws, 1, 26)
    _set_col_width(ws, 2, 12)
    _set_col_width(ws, 3, 12)
    _set_col_width(ws, 4, 14)
    _set_col_width(ws, 5, 16)
    _set_col_width(ws, 6, 16)
    _set_col_width(ws, 7, 16)
    _set_col_width(ws, 8, 18)
    _set_col_width(ws, 9, 16)
    _set_col_width(ws, 10, 16)
    _set_col_width(ws, 11, 18)

    _title_cell(ws, 1, "SteadyPlan — Performance Report", 11)
    cell = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y at %H:%M')}")
    cell.font = _SUBTITLE_FONT

    _header_row(ws, 4, [
        "Entity",
        "Start",
        "End",
        "Months",
        "Total Return",
        "Annualised",
        "Contributed",
        "Opening / Imported",
        "Gain / Interest",
        "Vs Plan",
        "Current Value",
    ])

    GBP = '£#,##0.00'
    PCT = '0.00"%"'
    row = 5

    def _performance_export_account_context(acc):
        wrapper = (acc.get("wrapper_type") or "").strip().lower()
        if wrapper == "cash isa":
            rate = to_float(acc.get("cash_interest_rate") or 0)
            return {
                "rate": rate,
                "subtitle": f"Cash interest rate: {rate*100:.1f}%",
                "gain_label": "Interest / Cash gain",
            }
        if wrapper == "premium bonds":
            rate = to_float(acc.get("growth_rate_override") if acc.get("growth_rate_override") is not None else 0.033)
            return {
                "rate": rate,
                "subtitle": f"Expected prize rate: {rate*100:.1f}%",
                "gain_label": "Prize gain",
            }
        return {
            "rate": account_growth_rate(acc, assumptions),
            "subtitle": f"Assumed growth: {account_growth_rate(acc, assumptions)*100:.1f}%",
            "gain_label": "Market Gain / Loss",
        }

    def _append_summary(entity_name, perf):
        nonlocal row
        if not perf:
            _data_row(ws, row, [entity_name, "", "", 0, "", "", "", "", "", "", ""], bold=True)
            row += 1
            return
        labels = perf.get("labels") or []
        start_m = labels[0] if labels else ""
        end_m = labels[-1] if labels else ""
        _data_row(ws, row, [
            entity_name,
            start_m,
            end_m,
            int(perf.get("n_months") or 0),
            float(perf.get("total_return") or 0),
            float(perf.get("annualised_return") or 0) if perf.get("annualised_return") is not None else None,
            float(perf.get("total_contributed") or 0),
            float(perf.get("total_imported_baseline") or 0),
            float(perf.get("total_market_gain") or 0),
            float(perf.get("vs_plan") or 0),
            float(perf.get("current_value") or 0),
        ], bold=True, num_formats={5: PCT, 6: PCT, 7: GBP, 8: GBP, 9: GBP, 10: GBP, 11: GBP})
        row += 1

    if selected_account_id is None:
        _append_summary("Portfolio", perf_portfolio)

        for aid, payload in per_account_data.items():
            acc = account_map.get(aid)
            if not acc:
                continue
            _, rows = _filter_monthly_data_for_period(payload["rows"], selected_period)
            ctx = _performance_export_account_context(acc)
            assumed_monthly = to_float(acc.get("monthly_contribution", 0))
            perf_acc = compute_performance_series(rows, ctx["rate"], assumed_monthly)
            _append_summary(payload["account_name"], perf_acc)
    else:
        payload = per_account_data.get(selected_account_id, {"account_name": account_map[selected_account_id]["name"], "rows": []})
        acc = account_map[selected_account_id]
        _, rows = _filter_monthly_data_for_period(payload["rows"], selected_period)
        ctx = _performance_export_account_context(acc)
        assumed_monthly = to_float(acc.get("monthly_contribution", 0))
        perf_acc = compute_performance_series(rows, ctx["rate"], assumed_monthly)
        _append_summary(payload["account_name"], perf_acc)

    def _safe_sheet_title(base, used):
        s = (base or "Sheet").strip()
        s = "".join("-" if ch in (":", "\\", "/", "?", "*", "[", "]") else ch for ch in s)
        s = s.strip().strip("'")[:31]
        if not s:
            s = "Sheet"
        if s not in used:
            used.add(s)
            return s
        i = 2
        while True:
            suffix = f" {i}"
            candidate = (s[:31 - len(suffix)] + suffix)[:31]
            if candidate not in used:
                used.add(candidate)
                return candidate
            i += 1

    used_titles = {ws.title}

    def _add_detail_sheet(title, perf, subtitle=None, gain_label="Market Gain / Loss"):
        ws_d = wb.create_sheet(_safe_sheet_title(title, used_titles))
        _set_col_width(ws_d, 1, 10)
        _set_col_width(ws_d, 2, 16)
        _set_col_width(ws_d, 3, 16)
        _set_col_width(ws_d, 4, 18)
        _set_col_width(ws_d, 5, 18)
        _set_col_width(ws_d, 6, 16)
        _set_col_width(ws_d, 7, 12)

        _title_cell(ws_d, 1, f"SteadyPlan — {title}", 7)
        sub = ws_d.cell(row=2, column=1, value=subtitle or f"Assumed growth: {assumed_rate*100:.1f}%")
        sub.font = _SUBTITLE_FONT

        has_first_baseline_only = bool(
            perf
            and (perf.get("labels") or [])
            and len(perf.get("labels") or []) == 1
            and not perf.get("table_rows")
        )
        if has_first_baseline_only:
            ws_d.cell(row=4, column=1, value="First value saved").font = _DATA_FONT
            first_value = float((perf.get("actual_values") or [0])[-1] or 0)
            first_label = (perf.get("labels") or [""])[-1]
            ws_d.cell(
                row=5,
                column=1,
                value=f"First tracked value: £{first_value:,.2f} in {first_label}.",
            ).font = _DATA_FONT
            ws_d.cell(
                row=6,
                column=1,
                value="It is treated as an opening/imported baseline, not performance gain.",
            ).font = _DATA_FONT
            ws_d.cell(
                row=7,
                column=1,
                value="Complete next month's monthly update and the month-by-month table will appear.",
            ).font = _DATA_FONT
            return

        if not perf or not perf.get("table_rows"):
            ws_d.cell(row=4, column=1, value="Not enough data yet (need at least two monthly snapshots).").font = _DATA_FONT
            return

        _header_row(ws_d, 4, ["Month", "Opening", "Opening / Imported", "Contributions", gain_label, "Closing", "Return"])
        rows_chrono = list(reversed(perf["table_rows"]))
        for i, r in enumerate(rows_chrono, 5):
            _data_row(ws_d, i, [
                r["month_key"],
                float(r["opening"]),
                float(r.get("imported_baseline") or 0),
                float(r["contribution"]),
                float(r["market_gain"]),
                float(r["closing"]),
                float(r["return_pct"]),
            ], num_formats={2: GBP, 3: GBP, 4: GBP, 5: GBP, 6: GBP, 7: PCT})

    if selected_account_id is None:
        _add_detail_sheet("Portfolio (Monthly)", perf_portfolio, subtitle=f"Assumed growth: {assumed_rate*100:.1f}%", gain_label="Gain / Interest")

        for aid, payload in per_account_data.items():
            acc = account_map.get(aid)
            if not acc:
                continue
            _, rows = _filter_monthly_data_for_period(payload["rows"], selected_period)
            ctx = _performance_export_account_context(acc)
            assumed_monthly = to_float(acc.get("monthly_contribution", 0))
            perf_acc = compute_performance_series(rows, ctx["rate"], assumed_monthly)
            _add_detail_sheet(f"{payload['account_name']} (Monthly)", perf_acc, subtitle=ctx["subtitle"], gain_label=ctx["gain_label"])
    else:
        payload = per_account_data.get(selected_account_id, {"account_name": account_map[selected_account_id]["name"], "rows": []})
        acc = account_map[selected_account_id]
        _, rows = _filter_monthly_data_for_period(payload["rows"], selected_period)
        ctx = _performance_export_account_context(acc)
        assumed_monthly = to_float(acc.get("monthly_contribution", 0))
        perf_acc = compute_performance_series(rows, ctx["rate"], assumed_monthly)
        _add_detail_sheet(f"{payload['account_name']} (Monthly)", perf_acc, ctx["subtitle"], ctx["gain_label"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if selected_account_id is None:
        fname = f"performance_{date.today().isoformat()}.xlsx"
    else:
        safe = "".join(ch for ch in (account_map[selected_account_id]["name"] or "account") if ch.isalnum() or ch in (" ", "-", "_")).strip().replace(" ", "_")
        fname = f"performance_{safe}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
